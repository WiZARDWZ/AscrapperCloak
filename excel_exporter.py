from __future__ import annotations

import os
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import config
import db_layer
from monitor import export_area_excel

WORKSHEET_NAME = "Listings"
PRICE_NOT_FOUND_DISPLAY = "\u0646\u062a\u0648\u0627\u0646\u0633\u062a\u06cc\u0645 \u0642\u06cc\u0645\u062a \u0631\u0627 \u067e\u06cc\u062f\u0627 \u06a9\u0646\u06cc\u0645"
NORMAL_HEADERS = [
    "Area",
    "Address",
    "Property Type",
    "Beds",
    "Baths",
    "Parking",
    "Land Size",
    "Building Size",
    "Price",
    "Ad Price",
    "Inferred Range",
    "Agency",
    "Agents",
    "Inspection",
    "Auction",
    "Description",
    "URL",
]
DEBUG_HEADERS = [
    "AreaLabel", "SearchID", "ListingID", "ExternalID", "listing_id", "url", "address",
    "property_type", "bedrooms", "bathrooms", "parking",
    "LandSizeDisplay", "LandSizeSqm", "BuildingSizeDisplay", "BuildingSizeSqm", "FloorAreaDisplay", "FloorAreaSqm",
    "Price", "PriceStatus", "PriceSource", "AdPriceDisplay", "AdPriceLow", "AdPriceHigh",
    "InferredPriceRange", "InferredPriceLow", "InferredPriceHigh",
    "PriceInferenceStatus", "PriceInferenceLastError", "PriceInferenceLastAttemptAt",
    "PriceLow", "PriceHigh", "LastPriceCheck", "CurrentPriceDisplay", "EffectivePriceDisplay",
    "ListingLifecycleStatus", "StatusReason", "StatusEvidence", "NotFoundCount", "MissingCount",
    "FirstNotFoundAt", "LastNotFoundAt", "RemovedAt", "SoldAt", "LastStatusChangeAt",
    "agency_name", "agents", "inspection_short", "inspection_long", "auction_label", "auction_time",
    "description", "area_status", "area_last_seen_at", "scraped_at",
]


@dataclass
class ExcelExportResult:
    file_path: str
    generated_at: datetime
    active_listing_count: int
    area_label: str
    search_id: int | None = None
    area_id: int | None = None

    @property
    def filename(self) -> str:
        return os.path.basename(self.file_path)

    @property
    def rows_count(self) -> int:
        return self.active_listing_count

    @property
    def listing_count(self) -> int:
        return self.active_listing_count


def _connect():
    return db_layer.connect(config.DB_PATH)


def _money(value: Any) -> str | None:
    if value is None or value == "":
        return None
    try:
        amount = Decimal(str(value))
    except Exception:
        return None
    return f"${amount:,.0f}"


def effective_price_display(current_display: Any, inferred_low: Any = None, inferred_high: Any = None) -> str:
    display = str(current_display or "").strip()
    if display and display.lower() not in {"n/a", "na", "none", "null", "unknown"}:
        return display
    inferred = inferred_price_range(inferred_low, inferred_high)
    return inferred or PRICE_NOT_FOUND_DISPLAY


def effective_price_source(current_display: Any, inferred_low: Any = None, inferred_high: Any = None) -> str:
    display = str(current_display or "").strip()
    if display and display.lower() not in {"n/a", "na", "none", "null", "unknown"}:
        return "ad_price"
    if inferred_price_range(inferred_low, inferred_high):
        return "inferred_range"
    return "unknown"


def inferred_price_range(inferred_low: Any = None, inferred_high: Any = None) -> str:
    low_text = _money(inferred_low)
    high_text = _money(inferred_high)
    if low_text and high_text and low_text != high_text:
        return f"{low_text} - {high_text}"
    if low_text or high_text:
        return low_text or high_text or ""
    return ""


def get_user_export_areas(telegram_user_id: int) -> list[dict]:
    conn = _connect()
    try:
        return [row for row in db_layer.list_user_area_subscriptions(conn, int(telegram_user_id), active_only=True) if row.get("IsActive", True)]
    finally:
        conn.close()


def get_authorized_export_area(telegram_user_id: int, user_area_id: int) -> dict | None:
    for row in get_user_export_areas(int(telegram_user_id)):
        if int(row.get("UserAreaID") or 0) == int(user_area_id):
            return row
    return None



def current_setup_readiness(area: dict) -> dict:
    """Return whether the current active subscription/setup run is eligible for final Excel export."""
    readiness_keys = {"AreaSetupStatus", "BaselineStatus", "DetailBaselineStatus", "PriceBaselineStatus", "NotificationReadyAt"}
    if not any(key in area for key in readiness_keys):
        return {"ready": True, "reasons": ["legacy_area_without_setup_fields"]}
    area_status = str(area.get("AreaSetupStatus") or "").lower()
    baseline = str(area.get("BaselineStatus") or "pending").lower()
    detail = str(area.get("DetailBaselineStatus") or "pending").lower()
    price = str(area.get("PriceBaselineStatus") or "pending").lower()
    notification_ready = bool(area.get("NotificationReadyAt"))
    ready = (area_status == "ready" and baseline == "completed" and detail == "completed" and price in {"completed", "completed_with_unknowns", "skipped"} and notification_ready)
    reasons = []
    if area_status != "ready": reasons.append(f"setup_status_{area_status or 'missing'}")
    if baseline != "completed": reasons.append(f"module1_{baseline}")
    if detail != "completed": reasons.append(f"module3_{detail}")
    if price not in {"completed", "completed_with_unknowns", "skipped"}: reasons.append(f"module2_{price}")
    if not notification_ready: reasons.append("notification_ready_at_missing")
    return {"ready": ready, "reasons": reasons}


def setup_preparing_message(area: dict) -> str:
    label = area.get("AreaLabel") or "this area"
    return f"⏳ Setup is still preparing for {label}. Excel export will be available after the current baseline, detail, and price setup finishes."

def normalize_export_mode(mode: str | None = None) -> str:
    normalized = str(mode or "normal").strip().lower()
    return normalized if normalized in {"normal", "debug"} else "normal"


def _fetch_active_listing_rows(conn, search_id: int, area_label: str) -> list[dict]:
    cur = conn.cursor()
    cur.execute("SELECT NormalizedSearchURL FROM dbo.SuburbSearch WHERE SearchID=?", int(search_id))
    row = cur.fetchone()
    if not row:
        return []
    area_url = row[0]
    rows = db_layer.export_latest_to_rows(conn, area_url)
    for item in rows:
        item["AreaLabel"] = area_label
        item.setdefault("SearchID", int(search_id))
        item.setdefault("ListingID", item.get("internal_listing_id") or item.get("ListingID"))
        item.setdefault("ExternalID", item.get("listing_id") or item.get("ExternalID"))
        item.setdefault("CurrentPriceDisplay", item.get("price_display"))
        item.setdefault("InferredPriceLow", item.get("inferred_price_low"))
        item.setdefault("InferredPriceHigh", item.get("inferred_price_high"))
        item.setdefault("LandSizeDisplay", item.get("land_size_display"))
        item.setdefault("LandSizeSqm", item.get("land_size_sqm"))
        item.setdefault("BuildingSizeDisplay", item.get("building_size_display"))
        item.setdefault("BuildingSizeSqm", item.get("building_size_sqm"))
        item.setdefault("FloorAreaDisplay", item.get("floor_area_display"))
        item.setdefault("FloorAreaSqm", item.get("floor_area_sqm"))
        item.setdefault("EffectivePriceDisplay", item.get("Price") or effective_price_display(item.get("CurrentPriceDisplay"), item.get("InferredPriceLow"), item.get("InferredPriceHigh")))
        item.setdefault("AdPriceDisplay", item.get("CurrentPriceDisplay") or item.get("price_display") or item.get("detail_price_display"))
        item.setdefault("AdPriceLow", item.get("ad_price_low") if item.get("ad_price_low") is not None else item.get("price_low"))
        item.setdefault("AdPriceHigh", item.get("ad_price_high") if item.get("ad_price_high") is not None else item.get("price_high"))
        item.setdefault("InferredPriceRange", inferred_price_range(item.get("InferredPriceLow"), item.get("InferredPriceHigh")))
        item.setdefault("PriceInferenceStatus", item.get("PriceStatus") or item.get("price_inference_status"))
        item.setdefault("PriceInferenceLastError", item.get("price_inference_last_error"))
        item.setdefault("PriceInferenceLastAttemptAt", item.get("LastPriceCheck") or item.get("last_price_check"))
    return rows


def _inspection_display(row: dict) -> str:
    return row.get("inspection_short") or row.get("inspection_long") or row.get("inspection") or ""


def _auction_display(row: dict) -> str:
    label = row.get("auction_label")
    time_value = row.get("auction_time")
    if label and time_value:
        return f"{label} {time_value}"
    return label or time_value or row.get("auction") or ""


def _normal_row(row: dict) -> dict:
    ad_price = row.get("AdPriceDisplay") or row.get("CurrentPriceDisplay") or row.get("price_display") or ""
    inferred = row.get("InferredPriceRange") or inferred_price_range(row.get("InferredPriceLow"), row.get("InferredPriceHigh"))
    return {
        "Area": row.get("AreaLabel") or "",
        "Address": row.get("address") or "",
        "Property Type": row.get("property_type") or "",
        "Beds": row.get("bedrooms") or "",
        "Baths": row.get("bathrooms") or "",
        "Parking": row.get("parking") or "",
        "Land Size": row.get("LandSizeDisplay") or row.get("land_size_display") or "",
        "Building Size": row.get("BuildingSizeDisplay") or row.get("building_size_display") or "",
        "Price": effective_price_display(ad_price, row.get("InferredPriceLow"), row.get("InferredPriceHigh")),
        "Ad Price": ad_price,
        "Inferred Range": inferred,
        "Agency": row.get("agency_name") or row.get("agency") or "",
        "Agents": row.get("agents") or "",
        "Inspection": _inspection_display(row),
        "Auction": _auction_display(row),
        "Description": row.get("description") or "",
        "URL": row.get("url") or "",
    }


def _debug_row(row: dict) -> dict:
    normalized = dict(row)
    normalized.setdefault("Price", row.get("Price") or effective_price_display(row.get("CurrentPriceDisplay") or row.get("price_display"), row.get("InferredPriceLow") or row.get("inferred_price_low"), row.get("InferredPriceHigh") or row.get("inferred_price_high")))
    normalized.setdefault("PriceStatus", row.get("PriceStatus") or row.get("price_inference_status") or "unknown_pending_retry")
    normalized.setdefault("PriceSource", row.get("PriceSource") or effective_price_source(row.get("AdPriceDisplay") or row.get("CurrentPriceDisplay") or row.get("price_display"), row.get("InferredPriceLow") or row.get("inferred_price_low"), row.get("InferredPriceHigh") or row.get("inferred_price_high")))
    normalized.setdefault("AdPriceDisplay", row.get("AdPriceDisplay") or row.get("CurrentPriceDisplay") or row.get("price_display") or row.get("detail_price_display") or "")
    normalized.setdefault("AdPriceLow", row.get("AdPriceLow") if row.get("AdPriceLow") is not None else row.get("ad_price_low"))
    normalized.setdefault("AdPriceHigh", row.get("AdPriceHigh") if row.get("AdPriceHigh") is not None else row.get("ad_price_high"))
    normalized.setdefault("InferredPriceLow", row.get("InferredPriceLow") or row.get("inferred_price_low"))
    normalized.setdefault("InferredPriceHigh", row.get("InferredPriceHigh") or row.get("inferred_price_high"))
    normalized.setdefault("InferredPriceRange", row.get("InferredPriceRange") or inferred_price_range(normalized.get("InferredPriceLow"), normalized.get("InferredPriceHigh")))
    normalized.setdefault("PriceInferenceStatus", row.get("PriceInferenceStatus") or normalized.get("PriceStatus"))
    normalized.setdefault("PriceInferenceLastError", row.get("PriceInferenceLastError") or row.get("price_inference_last_error") or "")
    normalized.setdefault("PriceInferenceLastAttemptAt", row.get("PriceInferenceLastAttemptAt") or row.get("LastPriceCheck") or row.get("last_price_check") or "")
    normalized.setdefault("ListingLifecycleStatus", row.get("ListingLifecycleStatus") or row.get("area_status") or "active")
    normalized.setdefault("MissingCount", row.get("MissingCount") if row.get("MissingCount") is not None else row.get("NotFoundCount"))
    return normalized


def _apply_normal_format(ws) -> None:
    widths = {
        "Area": 24, "Address": 45, "Property Type": 18, "Beds": 10, "Baths": 10, "Parking": 10,
        "Land Size": 16, "Building Size": 18, "Price": 24, "Ad Price": 24, "Inferred Range": 26,
        "Agency": 28, "Agents": 35, "Inspection": 28, "Auction": 28, "Description": 80, "URL": 55,
    }
    wrap_headers = {"Description", "Agents", "Inspection", "Auction", "URL"}
    for idx, header in enumerate(NORMAL_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header, 16)
        if header in wrap_headers:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=idx).alignment = Alignment(wrap_text=True, vertical="top")


def get_zero_row_diagnostics(telegram_user_id: int | None, user_area_id: int | None, search_id: int | None) -> dict:
    conn = _connect()
    try:
        return db_layer.get_excel_export_zero_row_diagnostics(conn, telegram_user_id, user_area_id, search_id)
    finally:
        conn.close()


def build_active_listings_excel(search_id: int, area_label: str, output_dir: str | None = None, mode: str = "normal") -> ExcelExportResult:
    mode = normalize_export_mode(mode)
    out_dir = output_dir or config.OUTPUT_DIR
    os.makedirs(out_dir, exist_ok=True)
    conn = _connect()
    try:
        rows = _fetch_active_listing_rows(conn, int(search_id), area_label)
    finally:
        conn.close()

    headers = NORMAL_HEADERS if mode == "normal" else DEBUG_HEADERS
    wb = Workbook()
    ws = wb.active
    ws.title = WORKSHEET_NAME
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    export_rows = rows if mode == "debug" else [row for row in rows if str(row.get("ListingLifecycleStatus") or row.get("area_status") or "active").strip().lower() == "active"]
    for row in export_rows:
        normalized = _normal_row(row) if mode == "normal" else _debug_row(row)
        ws.append([normalized.get(header, "") for header in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    if mode == "normal":
        _apply_normal_format(ws)
    generated_at = datetime.now()
    safe_label = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in str(area_label))[:60]
    file_path = os.path.join(out_dir, f"{safe_label}_{mode}_{generated_at.strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(file_path)
    return ExcelExportResult(file_path=file_path, generated_at=generated_at, active_listing_count=len(export_rows), area_label=str(area_label), search_id=int(search_id), area_id=int(search_id))


__all__ = [
    "ExcelExportResult",
    "DEBUG_HEADERS",
    "NORMAL_HEADERS",
    "PRICE_NOT_FOUND_DISPLAY",
    "WORKSHEET_NAME",
    "build_active_listings_excel",
    "effective_price_display",
    "inferred_price_range",
    "normalize_export_mode",
    "export_area_excel",
    "get_authorized_export_area",
    "get_user_export_areas",
    "get_zero_row_diagnostics",
    "current_setup_readiness",
    "setup_preparing_message",
]
