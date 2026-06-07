import os
import sys
from decimal import Decimal

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from openpyxl import load_workbook

import excel_exporter


def test_effective_price_prefers_display_then_inferred_range():
    assert excel_exporter.effective_price_display("Contact agent", Decimal("900000"), Decimal("950000")) == "Contact agent"
    assert excel_exporter.effective_price_display("", Decimal("900000"), Decimal("950000")) == "$900,000 - $950,000"
    assert excel_exporter.effective_price_display(None, Decimal("900000"), Decimal("900000")) == "$900,000"


def test_get_user_export_areas_returns_active_user_subscriptions(monkeypatch):
    class Conn:
        def close(self):
            pass

    rows = [
        {"UserAreaID": 1, "TelegramUserID": 7, "SearchID": 11, "AreaLabel": "Active", "IsActive": True},
        {"UserAreaID": 2, "TelegramUserID": 7, "SearchID": 12, "AreaLabel": "Inactive", "IsActive": False},
    ]
    monkeypatch.setattr(excel_exporter, "_connect", lambda: Conn())
    monkeypatch.setattr(excel_exporter.db_layer, "list_user_area_subscriptions", lambda conn, user_id, active_only=True: rows)

    areas = excel_exporter.get_user_export_areas(7)

    assert [area["UserAreaID"] for area in areas] == [1]


def test_build_active_listings_excel_creates_workbook_with_effective_price(monkeypatch, tmp_path):
    class Conn:
        def close(self):
            pass

    rows = [
        {
            "AreaLabel": "Tanglewood, NSW 2488",
            "SearchID": 22,
            "ListingID": 101,
            "ExternalID": "rea-101",
            "Address": "1 Test Street",
            "Suburb": "Tanglewood",
            "State": "NSW",
            "Postcode": "2488",
            "PropertyType": "House",
            "Bedrooms": 3,
            "Bathrooms": 2,
            "CarSpaces": 1,
            "LandSize": 500,
            "CurrentStatus": "active",
            "CurrentPriceDisplay": "Contact agent",
            "CurrentNumericPrice": None,
            "InferredPriceLow": Decimal("900000"),
            "InferredPriceHigh": Decimal("950000"),
            "EffectivePriceDisplay": excel_exporter.effective_price_display("Contact agent", Decimal("900000"), Decimal("950000")),
            "PriceInferenceStatus": "completed",
            "LastPriceInferenceAt": None,
            "LastDetailRefreshAt": None,
            "ListingURL": "https://example.test/listing/101",
            "Agent": "A Agent",
            "Agency": "A Agency",
            "FirstSeenAt": None,
            "LastSeenAt": None,
            "LastUpdatedAt": None,
            "SnapshotTimestamp": None,
        }
    ]
    monkeypatch.setattr(excel_exporter, "_connect", lambda: Conn())
    monkeypatch.setattr(excel_exporter, "_fetch_active_listing_rows", lambda conn, search_id, area_label: rows)

    result = excel_exporter.build_active_listings_excel(22, "Tanglewood, NSW 2488", output_dir=str(tmp_path))

    assert result.active_listing_count == 1
    assert result.file_path.endswith(".xlsx")
    wb = load_workbook(result.file_path)
    ws = wb[excel_exporter.WORKSHEET_NAME]
    headers = [cell.value for cell in ws[1]]
    values = [cell.value for cell in ws[2]]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
    assert headers.index("CurrentPriceDisplay") != headers.index("InferredPriceLow")
    assert values[headers.index("CurrentPriceDisplay")] == "Contact agent"
    assert values[headers.index("EffectivePriceDisplay")] == "Contact agent"
