import argparse
import json
import os
import tempfile
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional

import config
import db_layer
import module1_list_scraper
import module2_infer_prices
import module3_enrich_details
from realestate_errors import RealEstateBlockedError
from db_layer import (
    activate_area_subscriptions,
    connect,
    ensure_sort_list_date,
    export_latest_to_rows,
    get_listing_internal_id_by_external_id,
    get_or_create_area,
    get_unsent_events_for_area as db_get_unsent_events_for_area,
    ingest_full_rows,
    init_db,
    listing_seen_in_area,
    mark_events_sent,
    parse_price_bounds_from_text,
    parse_price_range,
    upsert_area_monitoring_state,
    upsert_price_inference_state,
)
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

WATCH_EVENT_TYPES = [
    "new_listing",
    "back_on_market",
    "price_change",
    "agent_change",
    "inspection_or_auction_change",
    "removed_or_missing",
]


def _module3_retryable_last_result() -> Dict[str, Any] | None:
    result = getattr(module3_enrich_details.module3_run, "last_result", {}) or {}
    status = str(result.get("status") or "").lower()
    if status.startswith("retry_wait"):
        return result
    return None




def _module3_output_is_reliable(rows: List[Dict[str, Any]]) -> bool:
    result = getattr(module3_enrich_details.module3_run, "last_result", {}) or {}
    status = str(result.get("status") or "").lower()
    if status.startswith("retry_wait"):
        return False
    success_count = result.get("success_count")
    if success_count is not None:
        try:
            return int(success_count) > 0
        except (TypeError, ValueError):
            return False
    return any(str(row.get("detail_scraped_at") or row.get("detail_status") or row.get("description") or "").strip() for row in rows)


def _module2_retryable_last_result() -> Dict[str, Any] | None:
    result = getattr(module2_infer_prices.module2_run, "last_result", {}) or {}
    status = str(result.get("status") or "").lower()
    if status.startswith("retry_wait") or status in {"retry_wait_browser_recovery", "429_retry_wait"}:
        return result
    return None

def _area_slug(search_url: str) -> str:
    slug = search_url.replace("https://", "").replace("http://", "")
    slug = slug.replace("/", "_").replace("?", "_").replace("&", "_").replace("=", "_")
    return "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in slug)[:60]


def _direct_price_text(row: Dict[str, Any]) -> str | None:
    for key in ("price", "price_display", "detail_price_display"):
        text = str(row.get(key) or "").strip()
        if text and text.lower() not in {"n/a", "na", "none", "null", "unknown"}:
            return text
    return None


def _row_has_inferred_price(row: Dict[str, Any]) -> bool:
    return (
        row.get("price_inferred_low") is not None
        or row.get("price_inferred_high") is not None
        or row.get("InferredPriceLow") is not None
        or row.get("InferredPriceHigh") is not None
    )


def _module2_targets(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [row for row in rows if db_layer.listing_is_active_for_module2(row)]


def _merge_rows_by_listing_id(base_rows: List[Dict[str, Any]], overlay_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    by_id = {str(r.get("listing_id") or ""): r for r in overlay_rows if str(r.get("listing_id") or "").strip()}
    return [by_id.get(str(row.get("listing_id") or ""), row) for row in base_rows]


def _record_price_states_for_rows(db_path: str, area_id: int, rows: List[Dict[str, Any]], attempted_module2: bool) -> Dict[str, int]:
    counts = {"inferred": 0, "unknown": 0, "direct": 0}
    conn = connect(db_path)
    try:
        for row in rows:
            listing_id = str(row.get("listing_id") or "").strip()
            if not listing_id:
                continue
            direct = _direct_price_text(row)
            inferred_low = row.get("price_inferred_low") if row.get("price_inferred_low") is not None else row.get("InferredPriceLow")
            inferred_high = row.get("price_inferred_high") if row.get("price_inferred_high") is not None else row.get("InferredPriceHigh")
            method = row.get("price_method") or row.get("price_inferred_method")
            if _row_has_inferred_price(row):
                upsert_price_inference_state(conn, listing_id, area_id, "completed", inferred_low=inferred_low, inferred_high=inferred_high, method=method or "sliding_between_window")
                counts["inferred"] += 1
            elif attempted_module2:
                upsert_price_inference_state(
                    conn,
                    listing_id,
                    area_id,
                    "unknown_pending_retry",
                    last_error="price_not_inferred_after_sweep",
                    next_retry_at=datetime.now(),
                    method="sliding_between_window",
                )
                counts["unknown"] += 1
            elif direct:
                upsert_price_inference_state(conn, listing_id, area_id, "skipped_direct_price", inferred_low=None, inferred_high=None, method=method or "direct", increment_attempts=False)
                counts["direct"] += 1
        conn.commit()
    finally:
        conn.close()
    return counts


def ingest_single_listing_snapshot(
    db_path: str,
    area_url: str,
    row: Dict[str, Any],
    source: str = "new_listing_enrich",
    emit_events: bool = True,
) -> Dict[str, Any]:
    normalized_url = ensure_sort_list_date(area_url)
    run_id = ingest_full_rows(db_path, normalized_url, [row], full_scan=False, emit_events=emit_events)
    conn = connect(db_path)
    try:
        area_id = get_or_create_area(conn, normalized_url)
        listing_id = str(row.get("listing_id") or "").strip()
        if listing_id:
            if _direct_price_text(row) and not _row_has_inferred_price(row):
                upsert_price_inference_state(conn, listing_id, area_id, "skipped_direct_price", method=row.get("price_method") or "direct", increment_attempts=False)
            elif _row_has_inferred_price(row):
                inferred_low = row.get("price_inferred_low") if row.get("price_inferred_low") is not None else row.get("InferredPriceLow")
                inferred_high = row.get("price_inferred_high") if row.get("price_inferred_high") is not None else row.get("InferredPriceHigh")
                upsert_price_inference_state(conn, listing_id, area_id, "completed", inferred_low=inferred_low, inferred_high=inferred_high, method=row.get("price_inferred_method") or row.get("price_method") or "sliding_between_window")
            else:
                upsert_price_inference_state(conn, listing_id, area_id, "unknown_pending_retry", last_error="price_not_inferred_after_sweep", next_retry_at=datetime.now(), method="sliding_between_window")
        conn.commit()
    finally:
        conn.close()
    return {"run_id": run_id, "source": source, "listing_id": row.get("listing_id")}


def export_area_excel(db_path: str, area_url: str, output_dir: str) -> str:
    os.makedirs(output_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"

    conn = connect(db_path)
    rows = export_latest_to_rows(conn, area_url)
    conn.close()

    headers = [
        "Price", "PriceStatus", "PriceSource", "PriceLow", "PriceHigh", "LastPriceCheck",
        "listing_id", "url", "address", "property_type", "bedrooms", "bathrooms", "parking",
        "LandSizeDisplay", "LandSizeSqm", "BuildingSizeDisplay", "BuildingSizeSqm", "FloorAreaDisplay", "FloorAreaSqm",
        "effective_price_display", "price_display", "price_low", "price_high", "price_method", "inferred_price_low", "inferred_price_high",
        "inspection_short", "inspection_long", "auction_label", "auction_time",
        "agency_name", "agency_code", "agency_profile_url", "agency_address",
        "agents", "detail_price_display", "area_status", "area_last_seen_at", "scraped_at", "description",
    ]

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    datetime_columns = {"area_last_seen_at", "scraped_at"}
    for col_name in datetime_columns:
        col_idx = headers.index(col_name) + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(output_dir, f"{_area_slug(area_url)}_{ts}.xlsx")
    wb.save(out_path)
    return out_path


def light_check_area(search_url: str, pages: int = config.LIGHT_CHECK_PAGES, timeout: int = config.PIPELINE_TIMEOUT) -> Dict[str, Any]:
    init_db(config.DB_PATH)
    normalized_url = ensure_sort_list_date(search_url)
    rows1 = module1_list_scraper.scrape_search(normalized_url, max_pages=pages, timeout=timeout)

    conn = connect(config.DB_PATH)
    area_id = get_or_create_area(conn, normalized_url)
    new_listings: List[Dict[str, Any]] = []

    lightweight_rows = []
    for r in rows1:
        listing_id = str(r.get("listing_id") or "").strip()
        if not listing_id:
            continue

        if not listing_seen_in_area(conn, area_id, listing_id):
            new_listings.append(
                {
                    "listing_id": listing_id,
                    "url": r.get("url"),
                    "address": r.get("address"),
                    "bedrooms": r.get("bedrooms"),
                    "bathrooms": r.get("bathrooms"),
                    "parking": r.get("parking"),
                    "property_type": r.get("property_type"),
                    "inspection_short_label": r.get("inspection_short_label"),
                    "auction_label": r.get("auction_label"),
                    "price": r.get("price"),
                }
            )

        lightweight_rows.append(
            {
                "listing_id": listing_id,
                "url": r.get("url"),
                "address": r.get("address"),
                "property_type": r.get("property_type"),
                "bedrooms": r.get("bedrooms"),
                "bathrooms": r.get("bathrooms"),
                "parking": r.get("parking"),
                "price": r.get("price"),
                "inspection_short_label": r.get("inspection_short_label"),
                "inspection_long_label": r.get("inspection_long_label"),
                "auction_label": r.get("auction_label"),
                "auction_time": r.get("auction_time"),
            }
        )

    run_id = ingest_full_rows(config.DB_PATH, normalized_url, lightweight_rows, full_scan=False)
    conn.close()

    return {
        "new_listings": new_listings,
        "new_count": len(new_listings),
        "area_id": area_id,
        "run_id": run_id,
        "search_url": normalized_url,
    }


def enrich_single_listing(area_url: str, listing_id: str, listing_url: str, job_id: Optional[int] = None, perf_mode: str = "normal") -> Dict[str, Any]:
    init_db(config.DB_PATH)
    normalized_url = ensure_sort_list_date(area_url)
    low_mode = perf_mode == "low"

    work_dir = config.OUTPUT_DIR
    if job_id is not None:
        work_dir = os.path.join(config.OUTPUT_DIR, f"job_{job_id}")
    os.makedirs(work_dir, exist_ok=True)

    temp_paths: List[str] = []
    base_row = [{"listing_id": listing_id, "url": listing_url, "price": "N/A"}]
    with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False, encoding="utf-8") as f:
        json.dump(base_row, f, ensure_ascii=False)
        single_input = f.name
    temp_paths.append(single_input)

    try:
        _, json3 = module3_enrich_details.module3_run(
            area_search_url=normalized_url,
            input_file=single_input,
            out_dir=work_dir,
            only_if_missing=False,
            wait_timeout=config.MODULE3_WAIT_TIMEOUT,
            sleep_between=config.MODULE3_SLEEP_BETWEEN * (2 if low_mode else 1),
            empty_retry=config.MODULE3_EMPTY_RETRY,
        )
        if not json3:
            raise RuntimeError("module3 failed for single listing")

        with open(json3, "r", encoding="utf-8") as f:
            enriched_rows = json.load(f)
        row = enriched_rows[0] if enriched_rows else {"listing_id": listing_id, "url": listing_url}

        srp_price = str(row.get("price") or "").strip()
        pdp_price = str(row.get("detail_price_display") or "").strip()

        price_text = srp_price if srp_price and srp_price.upper() != "N/A" else pdp_price
        low, high = parse_price_bounds_from_text(price_text)
        if price_text and price_text.upper() != "N/A":
            row["price"] = price_text
            row["AdPriceDisplay"] = price_text
            row["AdPriceLow"] = low
            row["AdPriceHigh"] = high
            row["PriceSource"] = "ad_price"
            row["price_method"] = "direct_from_pdp" if price_text == pdp_price else "direct"
        else:
            with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False, encoding="utf-8") as f:
                json.dump([row], f, ensure_ascii=False)
                input_for_m2 = f.name
            temp_paths.append(input_for_m2)
            _, json2 = module2_infer_prices.module2_run(
                base_list_url=normalized_url,
                input_file=input_for_m2,
                out_dir=work_dir,
                window_width=config.MODULE2_WINDOW_WIDTH,
                step=config.MODULE2_STEP,
                max_high=config.MODULE2_MAX_HIGH,
                max_pages_per_window=config.MODULE2_MAX_PAGES_PER_WINDOW,
                only_overwrite_na=True,
                smart_start=True,
            )
            if json2:
                with open(json2, "r", encoding="utf-8") as f:
                    rows2 = json.load(f)
                if rows2:
                    row = rows2[0]

        result = {
            "listing_id": row.get("listing_id") or listing_id,
            "url": row.get("url") or listing_url,
            "address": row.get("address"),
            "price_display": row.get("price") or row.get("detail_price_display"),
            "ad_price_display": row.get("AdPriceDisplay") or row.get("price") or row.get("detail_price_display"),
            "ad_price_low": row.get("AdPriceLow"),
            "ad_price_high": row.get("AdPriceHigh"),
            "inferred_price_low": row.get("price_inferred_low") if row.get("price_inferred_low") is not None else row.get("InferredPriceLow"),
            "inferred_price_high": row.get("price_inferred_high") if row.get("price_inferred_high") is not None else row.get("InferredPriceHigh"),
            "price_method": row.get("price_method"),
            "inspection_short_label": row.get("inspection_short_label"),
            "inspection_long_label": row.get("inspection_long_label"),
            "auction_label": row.get("auction_label"),
            "auction_time": row.get("auction_time"),
            "agency_name": row.get("agency_name") or row.get("agency"),
            "agency_code": row.get("agency_code"),
            "agents": row.get("agents") or [],
            "description": row.get("description"),
            "land_size_display": row.get("LandSizeDisplay") or row.get("land_size_display"),
            "land_size_sqm": row.get("LandSizeSqm") or row.get("land_size_sqm"),
            "building_size_display": row.get("BuildingSizeDisplay") or row.get("building_size_display"),
            "building_size_sqm": row.get("BuildingSizeSqm") or row.get("building_size_sqm"),
            "floor_area_display": row.get("FloorAreaDisplay") or row.get("floor_area_display"),
            "floor_area_sqm": row.get("FloorAreaSqm") or row.get("floor_area_sqm"),
            "property_type": row.get("property_type"),
            "bedrooms": row.get("bedrooms"),
            "bathrooms": row.get("bathrooms"),
            "parking": row.get("parking"),
        }
        row.setdefault("listing_id", result["listing_id"])
        row.setdefault("url", result["url"])
        row.setdefault("price", result["price_display"])
        persist_result = ingest_single_listing_snapshot(config.DB_PATH, normalized_url, row, source="new_listing_enrich", emit_events=True)
        result["persist_result"] = persist_result
        return result
    finally:
        for path in temp_paths:
            try:
                if path and os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass


def full_run_area(
    search_url: str,
    on_progress: Optional[Callable[[str, Optional[Dict[str, Any]]], None]] = None,
    cancel_token: Any = None,
    on_log: Optional[Callable[[str], None]] = None,
    perf_mode: Optional[str] = None,
) -> Dict[str, Any]:
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    init_db(config.DB_PATH)
    normalized_url = ensure_sort_list_date(search_url)
    # TODO: Low Mode is intentionally disabled for now. Re-enable in a later performance profile refactor.

    if on_progress:
        on_progress("module1_start", None)
    rows1 = module1_list_scraper.scrape_search(
        normalized_url,
        max_pages=config.MAX_PAGES_MODULE1,
        timeout=config.PIPELINE_TIMEOUT,
        cancel_token=cancel_token,
        on_log=on_log,
        on_progress=on_progress,
    )
    if getattr(cancel_token, "is_set", lambda: False)():
        return {"status": "cancelled", "area_id": None, "run_id": None, "excel_path": None, "events_count": 0}
    rows_module1 = len(rows1)
    if on_progress:
        on_progress("module1_done", {"rows": rows_module1})
    if rows_module1 == 0:
        module1_state = getattr(module1_list_scraper.scrape_search, "last_result", {}) or {}
        if module1_state.get("status") == "no_results" or module1_state.get("stop_reason") == "no_results":
            conn = connect(config.DB_PATH)
            try:
                area_id = get_or_create_area(conn, normalized_url)
            finally:
                conn.close()
            return {
                "status": "success",
                "run_id": None,
                "excel_path": None,
                "events_count": 0,
                "area_id": area_id,
                "rows_module1": 0,
                "rows_full": 0,
                "stop_reason": "no_results",
                "page_state": "no_results",
            }
        raise RuntimeError(
            f"Module1 returned 0 rows without confirmed no_results. page_state={module1_state.get('page_state')} stop_reason={module1_state.get('stop_reason')}"
        )
    _, json1 = module1_list_scraper.save_results(rows1, out_dir=config.OUTPUT_DIR)

    if on_progress:
        on_progress("module3_start", None)
    _, json3 = module3_enrich_details.module3_run(
        area_search_url=normalized_url,
        input_file=json1,
        out_dir=config.OUTPUT_DIR,
        only_if_missing=True,
        wait_timeout=config.MODULE3_WAIT_TIMEOUT,
        sleep_between=config.MODULE3_SLEEP_BETWEEN,  # TODO: Low Mode is intentionally disabled for now. Re-enable in a later performance profile refactor.
        empty_retry=config.MODULE3_EMPTY_RETRY,
        cancel_token=cancel_token,
        on_progress=on_progress,
        on_log=on_log,
    )
    if getattr(cancel_token, "is_set", lambda: False)():
        return {"status": "cancelled", "area_id": None, "run_id": None, "excel_path": None, "events_count": 0}
    if not json3:
        retryable_module3 = _module3_retryable_last_result()
        if retryable_module3:
            raise RealEstateBlockedError(
                retryable_module3.get("reason") or "Module3 retryable browser/navigation interruption",
                retry_after_seconds=int(retryable_module3.get("retry_after_seconds") or getattr(config, "REA_RATE_LIMIT_BACKOFF_SECONDS", 21600)),
            )
        raise RuntimeError("Module3 failed to produce JSON output")
    if on_progress:
        on_progress("module3_done", None)

    with open(json3, "r", encoding="utf-8") as f:
        rows_after_module3 = json.load(f)
    if not _module3_output_is_reliable(rows_after_module3):
        retryable_module3 = _module3_retryable_last_result() or {"reason": "Module3 output unreliable", "retry_after_seconds": getattr(config, "REA_RATE_LIMIT_BACKOFF_SECONDS", 21600)}
        raise RealEstateBlockedError(
            retryable_module3.get("reason") or "Module3 output unreliable",
            retry_after_seconds=int(retryable_module3.get("retry_after_seconds") or getattr(config, "REA_RATE_LIMIT_BACKOFF_SECONDS", 21600)),
        )

    rows_for_module2 = _module2_targets(rows_after_module3)

    full_rows = rows_after_module3
    if rows_for_module2:
        if on_progress:
            on_progress("module2_start", {"rows": len(rows_for_module2)})
        with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False, encoding="utf-8") as f:
            json.dump(rows_for_module2, f, ensure_ascii=False)
            module2_input = f.name
        _, json2 = module2_infer_prices.module2_run(
            base_list_url=normalized_url,
            input_file=module2_input,
            out_dir=config.OUTPUT_DIR,
            window_width=config.MODULE2_WINDOW_WIDTH,
            step=config.MODULE2_STEP,
            max_high=config.MODULE2_MAX_HIGH,
            max_pages_per_window=config.MODULE2_MAX_PAGES_PER_WINDOW,
            only_overwrite_na=True,
            smart_start=True,
            cancel_token=cancel_token,
            on_log=on_log,
            on_progress=on_progress,
        )
        try:
            os.remove(module2_input)
        except Exception:
            pass

        if getattr(cancel_token, "is_set", lambda: False)():
            return {"status": "cancelled", "area_id": None, "run_id": None, "excel_path": None, "events_count": 0}
        if not json2:
            retryable_module2 = _module2_retryable_last_result()
            if retryable_module2:
                raise RealEstateBlockedError(
                    retryable_module2.get("stopped_reason") or retryable_module2.get("browser_recovery_action") or "Module2 retryable browser/navigation interruption",
                    retry_after_seconds=int(getattr(config, "REA_RATE_LIMIT_BACKOFF_SECONDS", 21600)),
                )
            raise RuntimeError("Module2 failed to produce JSON output")
        with open(json2, "r", encoding="utf-8") as f:
            rows_after_module2 = json.load(f)
        map_after_module2 = {str(r.get("listing_id") or ""): r for r in rows_after_module2}
        merged_rows = []
        for row in rows_after_module3:
            lid = str(row.get("listing_id") or "")
            merged_rows.append(map_after_module2.get(lid, row))
        full_rows = merged_rows
        if on_progress:
            on_progress("module2_done", {"rows": len(rows_for_module2)})

    rows_full = len(full_rows)
    run_id = ingest_full_rows(config.DB_PATH, normalized_url, full_rows, full_scan=True)

    conn = connect(config.DB_PATH)
    area_id = get_or_create_area(conn, normalized_url)
    events = db_get_unsent_events_for_area(conn, area_id, WATCH_EVENT_TYPES)
    events_count = len(events)
    conn.close()

    excel_path = export_area_excel(config.DB_PATH, normalized_url, config.OUTPUT_DIR)
    if on_progress:
        on_progress("export_done", None)
    return {
        "status": "success",
        "run_id": run_id,
        "excel_path": excel_path,
        "events_count": events_count,
        "area_id": area_id,
        "rows_module1": rows_module1,
        "rows_full": rows_full,
    }


def baseline_setup_area(
    search_url: str,
    on_progress: Optional[Callable[[str, Optional[Dict[str, Any]]], None]] = None,
    cancel_token: Any = None,
    on_log: Optional[Callable[[str], None]] = None,
    perf_mode: str = "normal",
) -> Dict[str, Any]:
    """Start one area setup by running Module1 only, then queue bounded setup phases."""
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    init_db(config.DB_PATH)
    normalized_url = ensure_sort_list_date(search_url)
    conn = connect(config.DB_PATH)
    try:
        area_id = get_or_create_area(conn, normalized_url)
        upsert_area_monitoring_state(conn, area_id, setup_status="preparing", module1_status="running", module3_status="pending", module2_status="pending", last_error=None, set_started=True)
        conn.commit()
    finally:
        conn.close()

    rows1 = module1_list_scraper.scrape_search(
        normalized_url,
        max_pages=config.INITIAL_BASELINE_MAX_PAGES,
        timeout=config.PIPELINE_TIMEOUT,
        cancel_token=cancel_token,
        on_log=on_log,
        on_progress=on_progress,
    )
    if getattr(cancel_token, "is_set", lambda: False)():
        return {"status": "cancelled", "area_id": area_id}
    if not rows1:
        module1_state = getattr(module1_list_scraper.scrape_search, "last_result", {}) or {}
        if module1_state.get("status") == "no_results" or module1_state.get("stop_reason") == "no_results":
            conn = connect(config.DB_PATH)
            try:
                upsert_area_monitoring_state(
                    conn,
                    area_id,
                    setup_status="ready",
                    module1_status="completed",
                    module3_status="completed",
                    module2_status="completed",
                    active_listing_count=0,
                    inferred_price_count=0,
                    unknown_price_count=0,
                    last_error=None,
                    set_ready=True,
                )
                activate_area_subscriptions(conn, area_id)
                conn.commit()
            finally:
                conn.close()
            return {
                "status": "ready",
                "area_id": area_id,
                "rows_module1": 0,
                "rows_full": 0,
                "active_listing_count": 0,
                "inferred_price_count": 0,
                "unknown_price_count": 0,
                "events_count": 0,
                "stop_reason": "no_results",
                "page_state": "no_results",
                "empty_market": True,
                "detail_batches_enqueued": 0,
                "price_setup_enqueued": False,
            }
        conn = connect(config.DB_PATH)
        try:
            last_error = f"Module1 returned 0 rows without confirmed no_results. page_state={module1_state.get('page_state')} stop_reason={module1_state.get('stop_reason')}"
            upsert_area_monitoring_state(conn, area_id, setup_status="failed", module1_status="failed", last_error=last_error)
            conn.commit()
        finally:
            conn.close()
        raise RuntimeError(last_error)

    # Persist Module1/listing-shell data only. Detail enrichment and price inference
    # are intentionally separate setup jobs so large suburbs never run as one
    # multi-hour baseline_setup_area job.
    try:
        run_id = ingest_full_rows(config.DB_PATH, normalized_url, rows1, full_scan=True, emit_events=False)
    except Exception as exc:
        conn = connect(config.DB_PATH)
        try:
            upsert_area_monitoring_state(
                conn,
                area_id,
                setup_status="failed",
                module1_status="completed",
                module3_status="pending",
                module2_status="pending",
                active_listing_count=len(rows1),
                last_error=f"failed_ingest: {config.mask_sensitive_text(exc)}",
            )
            conn.commit()
        finally:
            conn.close()
        raise

    batch_size = max(1, int(getattr(config, "BASELINE_DETAIL_BATCH_SIZE", 50)))
    planned_batches = (len(rows1) + batch_size - 1) // batch_size
    conn = connect(config.DB_PATH)
    try:
        db_layer.mark_search_baseline_completed(
            conn,
            area_id,
            listings_collected=len(rows1),
            new_count=0,
            pages_checked=None,
            total_pages_detected=None,
            stop_reason="module1_setup_batched",
        )
        upsert_area_monitoring_state(
            conn,
            area_id,
            setup_status="preparing",
            module1_status="completed",
            module3_status="pending",
            module2_status="pending",
            active_listing_count=len(rows1),
            inferred_price_count=0,
            unknown_price_count=0,
            last_error=f"details 0/{len(rows1)}",
        )
        detail_job = db_layer.enqueue_setup_detail_baseline_job(conn, area_id, dedupe_suffix="initial")
        conn.commit()
    finally:
        conn.close()
    if on_log:
        on_log(f"Baseline setup batched: area_id={area_id} rows_module1={len(rows1)} detail_batch_size={batch_size} detail_batches_planned={planned_batches}")
    return {
        "status": "setup_batched",
        "area_id": area_id,
        "run_id": run_id,
        "rows_module1": len(rows1),
        "active_listing_count": len(rows1),
        "detail_total": len(rows1),
        "detail_batch_size": batch_size,
        "detail_batches_planned": planned_batches,
        "detail_batches_enqueued": 1 if detail_job and detail_job.get("created", True) else 0,
        "detail_job": detail_job,
        "price_setup_enqueued": False,
    }


def run_monitor_for_area(search_url: str) -> Dict[str, Any]:
    return full_run_area(search_url)


def get_unsent_events_for_area(area_id: int, event_types: List[str] | None = None) -> List[Dict[str, Any]]:
    conn = connect(config.DB_PATH)
    try:
        return db_get_unsent_events_for_area(conn, area_id, event_types or WATCH_EVENT_TYPES)
    finally:
        conn.close()


def mark_events_sent_for_area(event_ids: List[int]) -> None:
    conn = connect(config.DB_PATH)
    try:
        mark_events_sent(conn, event_ids)
    finally:
        conn.close()


def _cli() -> None:
    parser = argparse.ArgumentParser(description="Run monitor (light/full) for one area URL")
    parser.add_argument("--url", required=True)
    parser.add_argument("--mode", choices=["full", "light"], default="full")
    parser.add_argument("--pages", type=int, default=1)
    args = parser.parse_args()

    if args.mode == "light":
        result = light_check_area(args.url, pages=max(1, args.pages), timeout=config.PIPELINE_TIMEOUT)
    else:
        result = full_run_area(args.url)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    _cli()
