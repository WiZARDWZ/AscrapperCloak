import json
import inspect
import os
import sqlite3
import sys
import tempfile
import types
import unittest
from datetime import datetime
from unittest import mock

sys.modules.setdefault("pyodbc", types.SimpleNamespace(connect=lambda *args, **kwargs: None))
sys.modules.setdefault("chrome_options_helper", types.SimpleNamespace(build_chrome_driver=lambda *a, **k: None, cleanup_chrome_driver=lambda *a, **k: None))
sys.modules.setdefault(
    "browser_recovery",
    types.SimpleNamespace(
        is_429_page=lambda *a, **k: False,
        raise_if_realestate_blocked=lambda *a, **k: None,
        recover_browser_after_429=lambda *a, **k: None,
        recover_browser_for_untrusted_state=lambda *a, **k: (None, 0, "rea_profile", "rotation_limit"),
        safe_driver_get=lambda driver, url, log_func=print: (True, None),
        is_retryable_navigation_error=lambda exc: "net::" in str(exc).lower() or "page.goto" in str(exc).lower(),
        BrowserSessionHealth=object,
        RecoveryPolicy=object,
        log_session_health=lambda *a, **k: None,
        safe_realestate_get_with_reset=lambda *a, **k: (True, None),
        reset_chrome_error_tab=lambda *a, **k: None,
        UNTRUSTED_RECOVERY_STATES=set(),
        same_session_kpsdk_recheck=lambda **kwargs: (kwargs.get("initial_result"), kwargs.get("initial_payload")),
    ),
)
try:
    import bs4  # noqa: F401
except Exception:
    sys.modules.setdefault("bs4", types.SimpleNamespace(BeautifulSoup=lambda *a, **k: None))
telegram_stub = types.ModuleType("telegram")
telegram_stub.InlineKeyboardButton = type("InlineKeyboardButton", (), {"__init__": lambda self, *a, **k: None})
telegram_stub.InlineKeyboardMarkup = type("InlineKeyboardMarkup", (), {"__init__": lambda self, *a, **k: None})
telegram_stub.KeyboardButton = type("KeyboardButton", (), {"__init__": lambda self, *a, **k: None})
telegram_stub.ReplyKeyboardMarkup = type("ReplyKeyboardMarkup", (), {"__init__": lambda self, *a, **k: None})
telegram_stub.Update = type("Update", (), {})
telegram_stub.Bot = type("Bot", (), {"__init__": lambda self, *a, **k: None, "send_message": mock.AsyncMock()})
sys.modules.setdefault("telegram", telegram_stub)
telegram_ext_stub = types.ModuleType("telegram.ext")
telegram_ext_stub.Application = types.SimpleNamespace(builder=lambda: types.SimpleNamespace(token=lambda _token: types.SimpleNamespace(build=lambda: types.SimpleNamespace(add_handler=lambda *a, **k: None, run_polling=lambda *a, **k: None))))
telegram_ext_stub.CallbackQueryHandler = type("CallbackQueryHandler", (), {"__init__": lambda self, *a, **k: None})
telegram_ext_stub.CommandHandler = type("CommandHandler", (), {"__init__": lambda self, *a, **k: None})
telegram_ext_stub.ContextTypes = types.SimpleNamespace(DEFAULT_TYPE=object)
telegram_ext_stub.MessageHandler = type("MessageHandler", (), {"__init__": lambda self, *a, **k: None})
telegram_ext_stub.filters = types.SimpleNamespace(TEXT=object(), COMMAND=object())
sys.modules.setdefault("telegram.ext", telegram_ext_stub)

import db_layer
import excel_exporter
import job_queue
import listing_change_detector
import listing_detail_refresher
import monitor
import monitoring_scheduler
import module1_list_scraper
import module2_infer_prices
import module3_enrich_details
import notification_formatter
import telegram_bot
from realestate_errors import RealEstateRateLimitedError
from area_parser import parse_area_to_sqm
from openpyxl import load_workbook


class DummyConn:
    def commit(self):
        pass

    def close(self):
        pass


class RecordingCursor:
    def __init__(self):
        self.executed_sql = []
        self.description = [("listing_id",)]

    def execute(self, sql, *params):
        self.executed_sql.append(str(sql))
        return self

    def fetchone(self):
        return None

    def fetchall(self):
        return []


class RecordingConn:
    def __init__(self):
        self.cursor_obj = RecordingCursor()

    def cursor(self):
        return self.cursor_obj


class LifecycleCursor:
    def __init__(self, conn):
        self.conn = conn
        self.rows = []
        self.description = None
        self.rowcount = 0

    def execute(self, sql, *params):
        sql_text = str(sql)
        self.conn.executed.append((sql_text, params))
        self.rows = []
        self.description = None
        self.rowcount = 0
        if "SELECT UserAreaID, SearchID, IsActive" in sql_text:
            user_id, user_area_id = int(params[0]), int(params[1])
            sub = next((row for row in self.conn.subscriptions if row["TelegramUserID"] == user_id and row["UserAreaID"] == user_area_id), None)
            self.rows = [(sub["UserAreaID"], sub["SearchID"], sub["IsActive"])] if sub else []
            self.description = [("UserAreaID",), ("SearchID",), ("IsActive",)]
        elif "UPDATE dbo.UserAreaSubscription" in sql_text and "SET IsActive=0" in sql_text:
            user_id, user_area_id = int(params[0]), int(params[1])
            for sub in self.conn.subscriptions:
                if sub["TelegramUserID"] == user_id and sub["UserAreaID"] == user_area_id and sub["IsActive"]:
                    sub["IsActive"] = 0
                    sub["SubscriptionStatus"] = "removed"
                    sub["NotifyEnabled"] = 0
                    self.rowcount = 1
        elif "SELECT COUNT(1)" in sql_text and "FROM dbo.UserAreaSubscription" in sql_text:
            search_id = int(params[0])
            count = sum(
                1 for sub in self.conn.subscriptions
                if sub["SearchID"] == search_id
                and sub["IsActive"]
                and sub.get("SubscriptionStatus", "active") in {"active", "preparing"}
            )
            self.rows = [(count,)]
            self.description = [("Count",)]
        elif "MERGE dbo.area_monitoring_state" in sql_text:
            area_id = int(params[0])
            setup_status = params[1]
            last_subscription_count = params[17]
            state = self.conn.area_states.setdefault(area_id, {})
            if setup_status is not None:
                state["setup_status"] = setup_status
            state["last_subscription_count"] = last_subscription_count
        elif "UPDATE dbo.user_area_subscription_state" in sql_text:
            area_id = int(params[-1])
            self.conn.subscription_state_updates.append(area_id)
        elif "UPDATE dbo.UserAreaSubscription" in sql_text and "SET NotifyEnabled=0" in sql_text:
            search_id = int(params[0])
            for sub in self.conn.subscriptions:
                if sub["SearchID"] == search_id:
                    sub["NotifyEnabled"] = 0
                    if sub.get("SubscriptionStatus") != "removed":
                        sub["SubscriptionStatus"] = "inactive"
        elif "UPDATE dbo.Job" in sql_text and "Status='cancelled'" in sql_text:
            search_id = int(params[1])
            for job in self.conn.jobs:
                if job["SearchID"] == search_id and job["Status"] in {"pending", "paused", "queued", "retry_wait", "scheduled"}:
                    job["Status"] = "cancelled"
                    job["LastError"] = params[0]
                    self.rowcount += 1
        return self

    def fetchone(self):
        return self.rows.pop(0) if self.rows else None

    def fetchall(self):
        rows, self.rows = self.rows, []
        return rows


class LifecycleConn:
    def __init__(self, subscriptions, jobs=None, area_states=None):
        self.subscriptions = subscriptions
        self.jobs = jobs or []
        self.area_states = area_states or {}
        self.subscription_state_updates = []
        self.executed = []
        self.commits = 0
        self.rollbacks = 0

    def cursor(self):
        return LifecycleCursor(self)

    def commit(self):
        self.commits += 1

    def rollback(self):
        self.rollbacks += 1

    def close(self):
        pass


class NoopCursor:
    description = [("value",)]

    def execute(self, *args, **kwargs):
        return self

    def fetchone(self):
        return None

    def fetchall(self):
        return []


class NoopConn:
    def __init__(self):
        self.cursor_obj = NoopCursor()

    def cursor(self):
        return self.cursor_obj

    def commit(self):
        pass

    def rollback(self):
        pass

    def close(self):
        pass


class MonitoringRefactorTests(unittest.TestCase):
    def _run_mocked_baseline(self, rows1, rows3, rows2=None):
        calls = []
        module2_kwargs = []
        rows2 = rows3 if rows2 is None else rows2
        with tempfile.TemporaryDirectory() as tmp:
            json1 = os.path.join(tmp, "m1.json")
            json3 = os.path.join(tmp, "m3.json")
            json2 = os.path.join(tmp, "m2.json")
            with open(json3, "w", encoding="utf-8") as f:
                json.dump(rows3, f)
            with open(json2, "w", encoding="utf-8") as f:
                json.dump(rows2, f)

            def scrape(*args, **kwargs):
                calls.append("module1")
                return rows1

            def save(rows, out_dir):
                with open(json1, "w", encoding="utf-8") as f:
                    json.dump(rows, f)
                return None, json1

            def module3_run(**kwargs):
                calls.append("module3")
                return None, json3

            def module2_run(**kwargs):
                calls.append("module2")
                module2_kwargs.append(kwargs)
                with open(kwargs["input_file"], "r", encoding="utf-8") as f:
                    module2_kwargs[-1]["input_rows"] = json.load(f)
                return None, json2

            patches = [
                mock.patch.object(monitor, "init_db", lambda _path: None),
                mock.patch.object(monitor, "connect", return_value=DummyConn()),
                mock.patch.object(monitor, "get_or_create_area", return_value=5),
                mock.patch.object(monitor, "upsert_area_monitoring_state"),
                mock.patch.object(monitor, "activate_area_subscriptions"),
                mock.patch.object(monitor, "ingest_full_rows", return_value=77),
                mock.patch.object(monitor, "upsert_price_inference_state"),
                mock.patch.object(monitor.module1_list_scraper, "scrape_search", side_effect=scrape),
                mock.patch.object(monitor.module1_list_scraper, "save_results", side_effect=save),
                mock.patch.object(monitor.module3_enrich_details, "module3_run", side_effect=module3_run),
                mock.patch.object(monitor.module2_infer_prices, "module2_run", side_effect=module2_run),
            ]
            with patches[0], patches[1], patches[2], patches[3] as state, patches[4], patches[5] as ingest, patches[6] as price_state, patches[7], patches[8], patches[9], patches[10] as module2_mock:
                result = monitor.baseline_setup_area("https://example.test/search")

        return {
            "calls": calls,
            "result": result,
            "module2_kwargs": module2_kwargs,
            "module2_call_count": module2_mock.call_count,
            "price_state": price_state,
            "ingest": ingest,
            "state": state,
        }

    def test_init_db_creates_monitoring_state_tables_idempotently(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "state.db")
            db_layer.init_db(path)
            db_layer.init_db(path)
            conn = sqlite3.connect(path)
            try:
                tables = {
                    row[0]
                    for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")
                }
            finally:
                conn.close()
        self.assertIn("area_monitoring_state", tables)
        self.assertIn("user_area_subscription_state", tables)
        self.assertIn("listing_price_inference_state", tables)

    def test_enqueue_baseline_setup_job_dedupes_active_area_job(self):
        conn = DummyConn()
        job_queue.enable_in_memory_store()
        try:
            with mock.patch.object(db_layer, "ensure_monitoring_state_tables", lambda _conn: None), \
                 mock.patch.object(db_layer, "get_area_monitoring_state", return_value={"setup_status": "preparing"}):
                first = db_layer.enqueue_baseline_setup_job(conn, 10, "https://example.test/search")
                second = db_layer.enqueue_baseline_setup_job(conn, 10, "https://example.test/search")
            self.assertTrue(first["created"])
            self.assertFalse(second["created"])
            self.assertTrue(second["duplicate"])
        finally:
            job_queue.disable_in_memory_store()

    def test_enqueue_baseline_setup_job_skips_ready_area(self):
        conn = DummyConn()
        with mock.patch.object(db_layer, "ensure_monitoring_state_tables", lambda _conn: None), \
             mock.patch.object(db_layer, "get_area_monitoring_state", return_value={"setup_status": "ready"}):
            result = db_layer.enqueue_baseline_setup_job(conn, 10, "https://example.test/search")
        self.assertEqual(result["reason"], "area_ready")
        self.assertFalse(result["created"])

    def test_baseline_module2_targets_all_active_listings(self):
        rows1 = [{"listing_id": str(i), "url": f"u{i}", "price": "$800,000" if i % 3 == 0 else "N/A"} for i in range(1, 60)]
        rows3 = [dict(row, address=f"A{row['listing_id']}") for row in rows1]
        outcome = self._run_mocked_baseline(rows1, rows3)

        self.assertEqual(outcome["calls"], ["module1", "module3", "module2"])
        self.assertEqual(outcome["module2_call_count"], 1)
        kwargs = outcome["module2_kwargs"][0]
        self.assertEqual(kwargs["target_mode"], "all")
        self.assertEqual(len(kwargs["target_listing_ids"]), 59)
        self.assertEqual(len(kwargs["input_rows"]), 59)
        self.assertEqual(outcome["result"]["module2_target_count"], 59)
        outcome["ingest"].assert_called_once()
        self.assertFalse(outcome["ingest"].call_args.kwargs["emit_events"])

    def test_baseline_module2_does_not_batch_by_ten(self):
        rows1 = [{"listing_id": str(i), "url": f"u{i}", "price": "N/A"} for i in range(1, 46)]
        rows3 = [dict(row, address=f"A{row['listing_id']}") for row in rows1]
        outcome = self._run_mocked_baseline(rows1, rows3)

        self.assertEqual(outcome["module2_call_count"], 1)
        self.assertEqual(len(outcome["module2_kwargs"][0]["target_listing_ids"]), 45)
        self.assertEqual(len(outcome["module2_kwargs"][0]["input_rows"]), 45)

    def test_direct_price_does_not_exclude_baseline_inference_target(self):
        rows = [{"listing_id": "direct-1", "url": "u1", "price": "$800,000"}]
        outcome = self._run_mocked_baseline(rows, [dict(rows[0], address="A")])

        self.assertEqual(outcome["module2_call_count"], 1)
        self.assertIn("direct-1", outcome["module2_kwargs"][0]["target_listing_ids"])

    def test_baseline_ready_when_some_module2_targets_remain_unknown(self):
        rows1 = [{"listing_id": str(i), "url": f"u{i}", "price": "$800,000" if i % 2 == 0 else "N/A"} for i in range(1, 60)]
        rows3 = [dict(row, address=f"A{row['listing_id']}") for row in rows1]
        inferred_rows = [
            dict(row, price_inferred_low=700000, price_inferred_high=750000, price_inferred_method="sliding_between_window")
            if int(row["listing_id"]) <= 40 else row
            for row in rows3
        ]
        outcome = self._run_mocked_baseline(rows1, rows3, rows2=inferred_rows)

        statuses = [call.args[3] for call in outcome["price_state"].call_args_list]
        self.assertEqual(statuses.count("completed"), 40)
        self.assertEqual(statuses.count("unknown_pending_retry"), 19)
        self.assertEqual(outcome["result"]["status"], "ready")
        self.assertEqual(outcome["result"]["unknown_price_count"], 19)

    def test_module2_render_states_are_retryable_interruptions(self):
        self.assertTrue(monitoring_scheduler._module2_interrupted({"status": "render_timeout"}))
        self.assertTrue(monitoring_scheduler._module2_interrupted({"status": "blank_render"}))
        self.assertTrue(monitoring_scheduler._module2_interrupted({"status": "unknown"}))

    def test_enrich_single_listing_perf_modes_do_not_raise_and_persist(self):
        row = {"listing_id": "42", "url": "https://example.test/42", "price": "$900,000", "address": "A"}
        with tempfile.TemporaryDirectory() as tmp:
            json3 = os.path.join(tmp, "m3.json")
            with open(json3, "w", encoding="utf-8") as f:
                json.dump([row], f)

            for mode in ("normal", "low"):
                with mock.patch.object(monitor, "init_db", lambda _path: None), \
                     mock.patch.object(monitor.module3_enrich_details, "module3_run", return_value=(None, json3)), \
                     mock.patch.object(monitor, "ingest_single_listing_snapshot", return_value={"run_id": 1}) as persist:
                    result = monitor.enrich_single_listing("https://example.test/search", "42", "https://example.test/42", perf_mode=mode)
                self.assertEqual(result["listing_id"], "42")
                persist.assert_called_once()

    def test_ingest_single_listing_snapshot_uses_full_scan_false(self):
        row = {"listing_id": "55", "url": "https://example.test/55", "price": "$1,000,000"}
        with mock.patch.object(monitor, "ingest_full_rows", return_value=9) as ingest, \
             mock.patch.object(monitor, "connect", return_value=DummyConn()), \
             mock.patch.object(monitor, "get_or_create_area", return_value=3), \
             mock.patch.object(monitor, "upsert_price_inference_state"):
            result = monitor.ingest_single_listing_snapshot("ignored.db", "https://example.test/search", row)
        self.assertEqual(result["run_id"], 9)
        self.assertFalse(ingest.call_args.kwargs["full_scan"])
        self.assertTrue(ingest.call_args.kwargs["emit_events"])

    def test_excel_price_priority(self):
        self.assertEqual(excel_exporter.effective_price_display("$1,000,000", 900000, 950000), "$1,000,000")
        self.assertEqual(excel_exporter.effective_price_display("", 900000, 950000), "$900,000 - $950,000")
        self.assertEqual(excel_exporter.effective_price_display("", None, None), excel_exporter.PRICE_NOT_FOUND_DISPLAY)

    def test_parse_area_to_sqm_square_metres(self):
        self.assertEqual(parse_area_to_sqm("301m²"), 301)
        self.assertEqual(parse_area_to_sqm("301 m²"), 301)
        self.assertEqual(parse_area_to_sqm("301sqm"), 301)
        self.assertEqual(parse_area_to_sqm("500m2"), 500)
        self.assertEqual(parse_area_to_sqm("1,200m²"), 1200)

    def test_parse_area_to_sqm_hectares(self):
        self.assertEqual(parse_area_to_sqm("8.09ha"), 80900)
        self.assertEqual(parse_area_to_sqm("1.2 ha"), 12000)
        self.assertEqual(parse_area_to_sqm("0.5 hectares"), 5000)

    def test_module3_extracts_301sqm_land_size(self):
        html = """<h1 class="property-info-address">21 Raspberry Crescent, Schofields, NSW 2762</h1>
        <ul class="property-info__primary-features" aria-label="House with 301m² land size with 4 bedrooms with study 2 bathrooms 2 car spaces">
          <li aria-label="4 bedrooms"><p>4</p></li><li aria-label="2 bathrooms"><p>2</p></li>
          <li aria-label="2 car spaces"><p>2</p></li><li aria-label="301m² land size"><p>301m²</p></li><p>House</p>
        </ul><span class="property-price property-info__price">For Sale $1,250,000 - $1,320,000</span>"""
        out = module3_enrich_details.extract_detail_data_from_html(html)
        self.assertEqual(out["address"], "21 Raspberry Crescent, Schofields, NSW 2762")
        self.assertEqual(out["property_type"], "House")
        self.assertEqual(out["bedrooms"], 4)
        self.assertEqual(out["bathrooms"], 2)
        self.assertEqual(out["parking"], 2)
        self.assertEqual(out["land_size_display"], "301m²")
        self.assertEqual(out["land_size_sqm"], 301)
        self.assertNotIn("building_size_display", out)
        self.assertEqual(out["detail_price_display"], "For Sale $1,250,000 - $1,320,000")

    def test_module3_extracts_hectare_land_size(self):
        html = """<h1 class="property-info-address">463, Moody Road, Cambrai, SA 5353</h1>
        <ul class="property-info__primary-features" aria-label="Other with 8.09ha land size">
          <li aria-label="8.09ha land size"><p>8.09ha</p></li><p>Other</p>
        </ul><span class="property-price property-info__price">$249,000 - $270,000</span>"""
        out = module3_enrich_details.extract_detail_data_from_html(html)
        self.assertEqual(out["land_size_display"], "8.09ha")
        self.assertEqual(out["land_size_sqm"], 80900)
        self.assertEqual(out["property_type"], "Other")
        self.assertEqual(out["detail_price_display"], "$249,000 - $270,000")
        self.assertNotIn("building_size_display", out)

    def test_module3_extracts_building_size(self):
        html = """<h1 class="property-info-address">50D Windsor Avenue, Magill, SA 5072</h1>
        <ul class="property-info__primary-features" aria-label="Townhouse with 132m² building size with 3 bedrooms 2 bathrooms 2 car spaces">
          <li aria-label="3 bedrooms"><p>3</p></li><li aria-label="2 bathrooms"><p>2</p></li>
          <li aria-label="2 car spaces"><p>2</p></li><li aria-label="132m² building size"><p>132m²</p></li><p>Townhouse</p>
        </ul><span class="property-price property-info__price">$860,000</span>"""
        out = module3_enrich_details.extract_detail_data_from_html(html)
        self.assertNotIn("land_size_display", out)
        self.assertEqual(out["building_size_display"], "132m²")
        self.assertEqual(out["building_size_sqm"], 132)
        self.assertEqual(out["bedrooms"], 3)
        self.assertEqual(out["bathrooms"], 2)
        self.assertEqual(out["parking"], 2)
        self.assertEqual(out["property_type"], "Townhouse")
        self.assertEqual(out["detail_price_display"], "$860,000")

    def test_module3_extracts_call_for_price_as_ad_price(self):
        html = '<span class="property-price property-info__price">Call for price</span>'
        out = module3_enrich_details.extract_detail_data_from_html(html)
        self.assertEqual(out["AdPriceDisplay"], "Call for price")
        self.assertEqual(out["Price"], "Call for price")
        self.assertEqual(out["PriceSource"], "ad_price")
        self.assertNotIn("AdPriceLow", out)
        self.assertNotIn("AdPriceHigh", out)

    def test_module3_extracts_land_size_with_comma(self):
        html = '<ul class="property-info__primary-features" aria-label="Residential land with 1,012m² land size"><li aria-label="1,012m² land size"><p>1,012m²</p></li><p>Residential land</p></ul>'
        out = module3_enrich_details.extract_detail_data_from_html(html)
        self.assertEqual(out["land_size_display"], "1,012m²")
        self.assertEqual(out["land_size_sqm"], 1012)

    def test_module3_extracts_full_mitchell_highway_detail(self):
        html = """
        <h1 class="property-info-address">35 Mitchell Highway, Coolabah, NSW 2831</h1>
        <ul class="property-info__primary-features"
            aria-label="Residential land with 1,012m² land size">
          <li aria-label="1,012m² land size">
            <p>1,012m²</p>
          </li>
          <p>Residential land</p>
        </ul>
        <span class="property-price property-info__price">Call for price</span>
        """
        out = module3_enrich_details.extract_detail_data_from_html(html)
        self.assertEqual(out["address"], "35 Mitchell Highway, Coolabah, NSW 2831")
        self.assertEqual(out["property_type"], "Residential land")
        self.assertEqual(out["land_size_display"], "1,012m²")
        self.assertEqual(out["land_size_sqm"], 1012)
        self.assertNotIn("building_size_display", out)
        self.assertEqual(out["detail_price_display"], "Call for price")
        self.assertEqual(out["AdPriceDisplay"], "Call for price")

    def test_module3_detects_sold_evidence(self):
        html = """<h1 class="property-info-address">1 Sold Street, Sydney NSW 2000</h1>
        <span class="property-price property-info__price">Sold prior to auction</span>
        <div>Sold prior to auction</div>"""
        out = module3_enrich_details.extract_detail_data_from_html(html)
        self.assertEqual(out["ListingLifecycleStatus"], "sold")
        self.assertEqual(out["StatusReason"], "sold_evidence")
        self.assertIn("Sold prior to auction", out["StatusEvidence"])


    def test_module3_active_detail_ignores_generic_sold_text(self):
        html = """
        <main>
          <h1 class="property-info-address">12 Active Road, Noona NSW 2835</h1>
          <span class="property-price property-info__price">Contact Agent</span>
          <script>{"marketing":"recently sold nearby homes"}</script>
          <section>For Sale by negotiation</section>
        </main>
        """
        out = module3_enrich_details.extract_detail_data_from_html(html)
        self.assertNotEqual(out.get("ListingLifecycleStatus"), "sold")
        self.assertNotEqual(out.get("status"), "sold")
        self.assertEqual(out.get("SoldEvidenceStrength"), "weak")

    def test_detail_refresh_weak_sold_recent_active_is_suppressed(self):
        old_state = {"external_id": "1", "listing_id": 9, "status": "active", "price_display": "Contact Agent"}
        row = {
            "listing_id": "1",
            "external_id": "1",
            "db_listing_id": 9,
            "url": "https://www.realestate.com.au/property-house-nsw-noona-1",
            "status": "sold",
            "ListingLifecycleStatus": "sold",
            "StatusReason": "weak_sold_evidence_ignored",
            "SoldEvidenceStrength": "weak",
            "detail_refresh_success": True,
        }
        with mock.patch.object(db_layer, "get_latest_listing_state", return_value=old_state), \
             mock.patch.object(db_layer, "_search_id_if_exists", return_value=4), \
             mock.patch.object(db_layer, "_recent_active_list_evidence", return_value={"recent_active": True}):
            result = db_layer.detect_and_record_changes_for_row(NoopConn(), "https://example.test/search", row, create_events=False)
        self.assertEqual(result["suppressed_sold_count"], 1)
        self.assertEqual(result["weak_sold_evidence_count"], 1)
        self.assertFalse(any(e["event_type"] in {"sold", "status_changed"} for e in result["events_detected"]))
        self.assertFalse(any(e["event_type"] in {"sold", "status_changed"} for e in result["should_notify_events"]))
        self.assertTrue(any(w.get("warning") == "suppressed_sold_due_to_recent_active_list_evidence" for w in result["warnings"]))

    def test_detail_refresh_strong_sold_allows_sold_events(self):
        old_state = {"external_id": "1", "listing_id": 9, "status": "active", "price_display": "Contact Agent"}
        row = {
            "listing_id": "1",
            "external_id": "1",
            "db_listing_id": 9,
            "url": "https://www.realestate.com.au/property-house-nsw-noona-1",
            "status": "sold",
            "ListingLifecycleStatus": "sold",
            "StatusReason": "sold_evidence",
            "StatusEvidence": "Sold on 1 Jan 2026",
            "SoldEvidenceStrength": "strong",
            "detail_refresh_success": True,
        }
        with mock.patch.object(db_layer, "get_latest_listing_state", return_value=old_state):
            result = db_layer.detect_and_record_changes_for_row(NoopConn(), "https://example.test/search", row, create_events=False)
        event_types = {e["event_type"] for e in result["events_detected"]}
        self.assertEqual(result["strong_sold_evidence_count"], 1)
        self.assertIn("status_changed", event_types)
        self.assertIn("sold", event_types)
        self.assertTrue(any(e.get("should_notify") for e in result["events_detected"] if e["event_type"] == "sold"))

    def test_module1_active_observation_restores_stale_sold_state(self):
        source = inspect.getsource(db_layer.ingest_full_rows)
        self.assertIn("CurrentStatus='active'", source)
        self.assertIn("ListingLifecycleStatus='active'", source)
        self.assertIn("Status='active'", source)
        self.assertIn("SoldAt=NULL", source)
        self.assertIn("RemovedAt=NULL", source)

    def test_detail_refresh_result_includes_sold_diagnostics(self):
        with mock.patch.object(db_layer, "connect", return_value=NoopConn()), \
             mock.patch.object(db_layer, "get_active_listings_for_detail_refresh", return_value=[{"listing_id": "1", "url": "u"}]), \
             mock.patch.object(listing_detail_refresher, "ENRICH_DETAIL_ROWS_FUNC", return_value=[{"listing_id": "1", "url": "u", "detail_refresh_success": True}]), \
             mock.patch.object(db_layer, "detect_and_record_changes_for_row", return_value={
                 "external_id": "1", "events_detected": [], "events_created": 0,
                 "should_notify_events": [], "warnings": [],
                 "suppressed_sold_count": 1, "weak_sold_evidence_count": 1, "strong_sold_evidence_count": 0,
             }):
            result = listing_detail_refresher.refresh_active_listings("https://example.test/search", dry_run=True)
        self.assertEqual(result["suppressed_sold_count"], 1)
        self.assertEqual(result["weak_sold_evidence_count"], 1)
        self.assertEqual(result["strong_sold_evidence_count"], 0)

    def test_module1_card_size_html_extraction(self):
        html = '<article><ul><li aria-label="301m² land size"><p>301m²</p></li></ul></article>'
        out = module1_list_scraper.extract_size_features_from_card_html(html)
        self.assertEqual(out["land_size_display"], "301m²")
        self.assertEqual(out["land_size_sqm"], 301)

    def test_normalize_listing_row_accepts_size_fields(self):
        row = db_layer.normalize_listing_row({
            "listing_id": "1",
            "url": "u1",
            "LandSizeDisplay": "301m²",
            "LandSizeSqm": 301,
            "BuildingSizeDisplay": "132m²",
            "BuildingSizeSqm": 132,
        })
        self.assertEqual(row["land_size_display"], "301m²")
        self.assertEqual(int(row["land_size_sqm"]), 301)
        self.assertEqual(row["building_size_display"], "132m²")
        self.assertEqual(int(row["building_size_sqm"]), 132)

    def test_detail_refresh_merge_persists_new_land_size(self):
        latest = {"external_id": "1", "listing_id": 9, "price_display": "$1", "property_type": "House"}
        row = db_layer.merge_enriched_listing_detail_with_latest({"listing_id": "1", "LandSizeDisplay": "1,012m²", "LandSizeSqm": 1012}, latest)
        normalized = db_layer.normalize_listing_row(row)
        self.assertEqual(normalized["land_size_display"], "1,012m²")
        self.assertEqual(int(normalized["land_size_sqm"]), 1012)

    def test_detail_refresh_merge_persists_new_call_for_price(self):
        latest = {"external_id": "1", "listing_id": 9, "property_type": "House"}
        row = db_layer.merge_enriched_listing_detail_with_latest({"listing_id": "1", "AdPriceDisplay": "Call for price"}, latest)
        normalized = db_layer.normalize_listing_row(row)
        self.assertEqual(normalized["price_display"], "Call for price")
        self.assertIsNone(normalized["price_low"])
        self.assertIsNone(normalized["price_high"])

    def test_detail_refresh_merge_does_not_clear_existing_size(self):
        latest = {"external_id": "1", "listing_id": 9, "land_size_display": "1,012m²", "land_size_sqm": 1012}
        row = db_layer.merge_enriched_listing_detail_with_latest({"listing_id": "1", "LandSizeDisplay": None}, latest)
        normalized = db_layer.normalize_listing_row(row)
        self.assertEqual(normalized["land_size_display"], "1,012m²")
        self.assertEqual(int(normalized["land_size_sqm"]), 1012)

    def test_newly_discovered_field_sends_notification_event(self):
        events = listing_change_detector.compare_listing_state(
            {"external_id": "1", "listing_id": 1, "land_size_display": None},
            {"external_id": "1", "listing_id": 1, "land_size_display": "1,012m²"},
        )
        event = next(e for e in events if e["field"] == "land_size_display")
        self.assertEqual(event["event_type"], "field_discovered")
        self.assertTrue(event["should_notify"])
        self.assertEqual(event["new_value"], "1,012m²")

    def test_real_size_change_sends_notification_event(self):
        events = listing_change_detector.compare_listing_state(
            {"external_id": "1", "listing_id": 1, "land_size_display": "1,000m²"},
            {"external_id": "1", "listing_id": 1, "land_size_display": "1,012m²"},
        )
        event = next(e for e in events if e["field"] == "land_size_display")
        self.assertEqual(event["event_type"], "size_changed")
        self.assertTrue(event["should_notify"])

    def test_price_label_discovery_and_change_notify(self):
        discovered = listing_change_detector.compare_listing_state(
            {"external_id": "1", "listing_id": 1, "price_display": None},
            {"external_id": "1", "listing_id": 1, "price_display": "Call for price"},
        )
        self.assertTrue(any(e["event_type"] == "field_discovered" and e["field"] == "AdPriceDisplay" and e["should_notify"] for e in discovered))
        changed = listing_change_detector.compare_listing_state(
            {"external_id": "1", "listing_id": 1, "price_display": "Contact Agent"},
            {"external_id": "1", "listing_id": 1, "price_display": "$860,000"},
        )
        self.assertTrue(any(e["event_type"] == "ad_price_changed" and e["should_notify"] for e in changed))

    def test_notification_formatter_renders_field_discovered(self):
        message = notification_formatter.format_notification_message({
            "EventType": "field_discovered",
            "OldValueJson": None,
            "NewValueJson": json.dumps("1,012m²"),
            "EventPayloadJson": json.dumps({"field": "land_size_display", "area_label": "Coolabah", "address": "35 Mitchell Highway", "price_display": "Call for price", "listing_url": "https://example.test/1"}),
        })
        self.assertIn("Land size added", message)
        self.assertIn("was: empty", message)
        self.assertIn("now: 1,012m²", message)
        self.assertIn("Call for price", message)

    def test_no_duplicate_event_if_field_unchanged(self):
        events = listing_change_detector.compare_listing_state(
            {"external_id": "1", "listing_id": 1, "land_size_display": "1,012m²"},
            {"external_id": "1", "listing_id": 1, "land_size_display": "1,012m²"},
        )
        self.assertFalse([e for e in events if e.get("field") == "land_size_display"])

    def test_listing_snapshot_insert_sql_matches_size_param_count(self):
        self.assertEqual(db_layer.LISTING_SNAPSHOT_INSERT_SQL.count("?"), 25)

    def test_ad_price_numeric_parsing_keeps_inferred_null(self):
        row = db_layer.normalize_listing_row({"listing_id": "1", "url": "u1", "price": "$850,000"})
        self.assertEqual(row["price_display"], "$850,000")
        self.assertEqual(int(row["price_low"]), 850000)
        self.assertEqual(int(row["price_high"]), 850000)

    def test_ad_price_non_numeric_has_no_numeric_bounds(self):
        row = db_layer.normalize_listing_row({"listing_id": "1", "url": "u1", "price": "Contact Agent"})
        self.assertEqual(row["price_display"], "Contact Agent")
        self.assertIsNone(row["price_low"])
        self.assertIsNone(row["price_high"])

    def test_module2_inferred_fields_separate_from_ad_price(self):
        row = db_layer.normalize_listing_row({
            "listing_id": "1",
            "url": "u1",
            "price": "$850,000",
            "price_inferred_low": 800000,
            "price_inferred_high": 850000,
        })
        self.assertEqual(row["price_display"], "$850,000")
        self.assertEqual(int(row["price_low"]), 850000)
        self.assertEqual(int(row["price_high"]), 850000)
        export_row = {
            "AdPriceDisplay": row["price_display"],
            "AdPriceLow": row["price_low"],
            "AdPriceHigh": row["price_high"],
            "InferredPriceLow": 800000,
            "InferredPriceHigh": 850000,
        }
        self.assertEqual(excel_exporter.inferred_price_range(export_row["InferredPriceLow"], export_row["InferredPriceHigh"]), "$800,000 - $850,000")

    def test_active_listing_under_offer_is_eligible_for_module2(self):
        rows = [
            {"listing_id": "1", "ListingLifecycleStatus": "active", "AdPriceDisplay": "Under Offer", "price": "Under Offer"},
            {"listing_id": "2", "ListingLifecycleStatus": "sold", "AdPriceDisplay": "Contact Agent"},
        ]
        targets = monitor._module2_targets(rows)
        self.assertEqual([row["listing_id"] for row in targets], ["1"])

    def test_lifecycle_sold_skips_module2_and_notifies_once(self):
        first = db_layer.next_listing_lifecycle_transition("active", 0, "sold")
        second = db_layer.next_listing_lifecycle_transition("sold", 0, "sold")
        self.assertEqual(first["new_status"], "sold")
        self.assertTrue(first["notify"])
        self.assertFalse(first["module2_eligible"])
        self.assertFalse(second["notify"])

    def test_lifecycle_first_not_found_does_not_notify_and_rechecks(self):
        transition = db_layer.next_listing_lifecycle_transition("active", 0, "not_found")
        self.assertEqual(transition["new_status"], "not_found")
        self.assertEqual(transition["not_found_count"], 1)
        self.assertFalse(transition["notify"])
        self.assertTrue(transition["enqueue_recheck"])
        self.assertFalse(transition["module2_eligible"])

    def test_active_to_not_found_enqueues_high_priority_recheck_job(self):
        job_queue.enable_in_memory_store()
        try:
            with mock.patch.object(db_layer, "ensure_listing_lifecycle_columns", lambda _conn: None), \
                 mock.patch.object(db_layer, "_one", side_effect=[("active", 0), ("ext-1", "https://example.test/1")]):
                transition = db_layer.apply_listing_lifecycle_signal(NoopConn(), 4, 9, "not_found", "missing_from_full_scan", "missing", create_event=False)
            self.assertEqual(transition["new_status"], "not_found")
            self.assertIsNotNone(transition["recheck_job"])
            self.assertEqual(transition["recheck_job"]["JobType"], job_queue.JOB_TYPE_LISTING_STATUS_RECHECK)
            self.assertEqual(transition["recheck_job"]["Priority"], job_queue.PRIORITY_LISTING_STATUS_RECHECK)
            self.assertEqual(json.loads(transition["recheck_job"]["PayloadJson"])["recheck_listing_id"], 9)
        finally:
            job_queue.disable_in_memory_store()

    def test_duplicate_not_found_recheck_job_is_deduped(self):
        job_queue.enable_in_memory_store()
        try:
            first = db_layer.enqueue_listing_status_recheck_job(NoopConn(), 4, 9, "ext-1", "https://example.test/1")
            second = db_layer.enqueue_listing_status_recheck_job(NoopConn(), 4, 9, "ext-1", "https://example.test/1")
            self.assertFalse(first.get("duplicate"))
            self.assertTrue(second.get("duplicate"))
            self.assertEqual(len(job_queue.get_active_jobs()), 1)
        finally:
            job_queue.disable_in_memory_store()

    def test_recheck_job_priority_runs_before_normal_refresh(self):
        job_queue.enable_in_memory_store()
        try:
            normal = job_queue.enqueue_job_once(job_queue.JOB_TYPE_DETAIL_REFRESH_EXISTING, search_id=4, priority=job_queue.PRIORITY_DETAIL_REFRESH)
            recheck = db_layer.enqueue_listing_status_recheck_job(NoopConn(), 4, 9, "ext-1", "https://example.test/1")
            claimed = job_queue.claim_next_job("worker")
            self.assertEqual(claimed["JobID"], recheck["JobID"])
            self.assertLess(recheck["Priority"], normal["Priority"])
        finally:
            job_queue.disable_in_memory_store()

    def test_lifecycle_second_not_found_becomes_removed(self):
        transition = db_layer.next_listing_lifecycle_transition("not_found", 1, "not_found")
        self.assertEqual(transition["new_status"], "removed")
        self.assertEqual(transition["event_type"], "removed")
        self.assertTrue(transition["notify"])
        self.assertFalse(transition["module2_eligible"])

    def test_lifecycle_not_found_recovery_resumes_module2(self):
        transition = db_layer.next_listing_lifecycle_transition("not_found", 1, "active")
        self.assertEqual(transition["new_status"], "active")
        self.assertEqual(transition["not_found_count"], 0)
        self.assertTrue(transition["module2_eligible"])
        self.assertFalse(transition["notify"])

    def test_lifecycle_technical_failure_keeps_previous_status(self):
        transition = db_layer.next_listing_lifecycle_transition("active", 0, "timeout")
        self.assertEqual(transition["new_status"], "active")
        self.assertFalse(transition["notify"])
        self.assertFalse(transition["enqueue_recheck"])

    def test_recheck_valid_listing_restores_active_and_resets_not_found(self):
        job = {"SearchID": 4, "UserAreaID": 7, "Payload": {"recheck_listing_id": 9, "listing_external_id": "ext-1"}}
        sub = {"UserAreaID": 7, "SearchID": 4, "SearchURL": "https://example.test/search"}
        with mock.patch.object(monitoring_scheduler, "_load_search_subscription", return_value=sub), \
             mock.patch("listing_detail_refresher.refresh_active_listings", return_value={"failed_count": 0, "errors": []}), \
             mock.patch.object(db_layer, "connect", return_value=NoopConn()), \
             mock.patch.object(db_layer, "_one", return_value=("active",)), \
             mock.patch.object(db_layer, "apply_listing_lifecycle_signal", return_value={"new_status": "active", "not_found_count": 0, "should_notify": False, "module2_eligible": True}) as apply_mock:
            result = monitoring_scheduler.run_listing_status_recheck_job(job, send_telegram=True)
        self.assertEqual(result["transition"]["new_status"], "active")
        apply_mock.assert_called_with(mock.ANY, 4, 9, "active", "status_recheck_valid", None, create_event=False)

    def test_recheck_second_not_found_marks_removed_and_notifies(self):
        job = {"SearchID": 4, "UserAreaID": 7, "Payload": {"recheck_listing_id": 9, "listing_external_id": "ext-1"}}
        sub = {"UserAreaID": 7, "SearchID": 4, "SearchURL": "https://example.test/search"}
        with mock.patch.object(monitoring_scheduler, "_load_search_subscription", return_value=sub), \
             mock.patch("listing_detail_refresher.refresh_active_listings", return_value={"failed_count": 1, "errors": [{"error": "page not found"}]}), \
             mock.patch.object(db_layer, "connect", return_value=NoopConn()), \
             mock.patch.object(db_layer, "apply_listing_lifecycle_signal", return_value={"new_status": "removed", "should_notify": True}) as apply_mock, \
             mock.patch.object(monitoring_scheduler, "_queue_notifications_for_search", return_value=[{"queued": 1}]) as notify_mock:
            result = monitoring_scheduler.run_listing_status_recheck_job(job, send_telegram=True)
        self.assertEqual(result["transition"]["new_status"], "removed")
        apply_mock.assert_called_with(mock.ANY, 4, 9, "not_found", "status_recheck_not_found", mock.ANY, create_event=True)
        notify_mock.assert_called_once_with(4, dry_run=False)

    def test_recheck_sold_evidence_marks_sold_and_notifies(self):
        job = {"SearchID": 4, "UserAreaID": 7, "Payload": {"recheck_listing_id": 9, "listing_external_id": "ext-1"}}
        sub = {"UserAreaID": 7, "SearchID": 4, "SearchURL": "https://example.test/search"}
        with mock.patch.object(monitoring_scheduler, "_load_search_subscription", return_value=sub), \
             mock.patch("listing_detail_refresher.refresh_active_listings", return_value={"failed_count": 0, "errors": []}), \
             mock.patch.object(db_layer, "connect", return_value=NoopConn()), \
             mock.patch.object(db_layer, "_one", return_value=("sold",)), \
             mock.patch.object(monitoring_scheduler, "_queue_notifications_for_search", return_value=[{"queued": 1}]) as notify_mock:
            result = monitoring_scheduler.run_listing_status_recheck_job(job, send_telegram=True)
        self.assertEqual(result["transition"]["new_status"], "sold")
        notify_mock.assert_called_once_with(4, dry_run=False)

    def test_recheck_technical_failure_does_not_mark_removed(self):
        job = {"SearchID": 4, "UserAreaID": 7, "Payload": {"recheck_listing_id": 9, "listing_external_id": "ext-1"}}
        sub = {"UserAreaID": 7, "SearchID": 4, "SearchURL": "https://example.test/search"}
        with mock.patch.object(monitoring_scheduler, "_load_search_subscription", return_value=sub), \
             mock.patch("listing_detail_refresher.refresh_active_listings", return_value={"failed_count": 1, "errors": [{"error": "timeout"}]}), \
             mock.patch.object(db_layer, "apply_listing_lifecycle_signal") as apply_mock:
            result = monitoring_scheduler.run_listing_status_recheck_job(job, send_telegram=True)
        self.assertEqual(result["status"], "retry_wait")
        apply_mock.assert_not_called()

    def test_inferred_range_change_notification_event(self):
        self.assertTrue(db_layer._price_range_materially_changed(200000, 250000, 250000, 300000))
        message = notification_formatter.format_notification_message({
            "EventType": "inferred_price_range_changed",
            "OldValueJson": json.dumps({"inferred_price_low": 200000, "inferred_price_high": 250000}),
            "NewValueJson": json.dumps({"inferred_price_low": 250000, "inferred_price_high": 300000}),
            "EventPayloadJson": json.dumps({"area_label": "A", "address": "B"}),
        })
        self.assertIn("Estimated range changed", message)
        self.assertIn("$200,000 - $250,000", message)
        self.assertIn("$250,000 - $300,000", message)

    def test_ad_and_inferred_price_changes_coalesce_to_one_listing_update(self):
        events = [
            {"EventID": 1, "EventType": "ad_price_changed", "SearchID": 4, "ListingID": 9, "RunID": 2, "OldValueJson": json.dumps({"price_display": "Contact Agent"}), "NewValueJson": json.dumps({"price_display": "Under Offer"}), "EventPayloadJson": json.dumps({"area_label": "A", "address": "B"})},
            {"EventID": 2, "EventType": "inferred_price_range_changed", "SearchID": 4, "ListingID": 9, "RunID": 2, "OldValueJson": json.dumps({"inferred_price_low": 200000, "inferred_price_high": 250000}), "NewValueJson": json.dumps({"inferred_price_low": 250000, "inferred_price_high": 300000}), "EventPayloadJson": json.dumps({"area_label": "A", "address": "B"})},
        ]
        coalesced = db_layer.coalesce_listing_price_update_events(events)
        self.assertEqual(len(coalesced), 1)
        self.assertEqual(coalesced[0]["EventType"], "listing_update")
        self.assertEqual(coalesced[0]["_absorbed_event_ids"], [2])
        message = notification_formatter.format_notification_message(coalesced[0])
        self.assertIn("Ad price changed", message)
        self.assertIn("Estimated range changed", message)

    def test_excel_export_result_has_area_label_and_required_price_columns(self):
        row = {
            "listing_id": "1",
            "url": "https://example.test/1",
            "Price": "$800,000",
            "PriceSource": "direct",
            "PriceStatus": "completed",
            "CurrentPriceDisplay": "$800,000",
            "InferredPriceLow": 700000,
            "InferredPriceHigh": 750000,
            "PriceInferenceLastError": "",
            "LastPriceCheck": datetime(2026, 1, 1, 12, 0, 0),
        }
        with tempfile.TemporaryDirectory() as tmp, \
             mock.patch.object(excel_exporter, "_connect", return_value=DummyConn()), \
             mock.patch.object(excel_exporter, "_fetch_active_listing_rows", return_value=[row]):
            result = excel_exporter.build_active_listings_excel(4, "Yadboro, NSW 2539", output_dir=tmp, mode="debug")
            workbook = load_workbook(result.file_path)
            headers = [cell.value for cell in workbook[excel_exporter.WORKSHEET_NAME][1]]

        self.assertEqual(result.area_label, "Yadboro, NSW 2539")
        self.assertEqual(result.search_id, 4)
        for header in [
            "Price",
            "PriceSource",
            "AdPriceDisplay",
            "AdPriceLow",
            "AdPriceHigh",
            "InferredPriceRange",
            "InferredPriceLow",
            "InferredPriceHigh",
            "PriceInferenceStatus",
            "PriceInferenceLastError",
            "PriceInferenceLastAttemptAt",
        ]:
            self.assertIn(header, headers)

    def test_export_latest_rows_does_not_filter_out_non_active_status_values(self):
        conn = RecordingConn()
        with mock.patch.object(db_layer, "ensure_monitoring_state_tables", lambda _conn: None), \
             mock.patch.object(db_layer, "_upsert_search", return_value=4):
            rows = db_layer.export_latest_to_rows(conn, "https://example.test/search")

        self.assertEqual(rows, [])
        combined_sql = "\n".join(conn.cursor_obj.executed_sql)
        self.assertNotIn("LOWER(COALESCE(lss.Status", combined_sql)
        self.assertNotIn("lss.Status, 'active'))='active'", combined_sql)

    def test_contact_agent_plus_inferred_range_price_priority(self):
        self.assertEqual(excel_exporter.effective_price_display("Contact Agent", 1000000, 1200000), "Contact Agent")
        self.assertEqual(excel_exporter.effective_price_source("Contact Agent", 1000000, 1200000), "ad_price")
        self.assertEqual(excel_exporter.inferred_price_range(1000000, 1200000), "$1,000,000 - $1,200,000")

    def test_no_ad_price_uses_inferred_range(self):
        self.assertEqual(excel_exporter.effective_price_display("", 1000000, 1200000), "$1,000,000 - $1,200,000")
        self.assertEqual(excel_exporter.effective_price_source("", 1000000, 1200000), "inferred_range")

    def test_no_ad_price_and_no_inferred_uses_unable_message(self):
        self.assertEqual(excel_exporter.effective_price_display("", None, None), excel_exporter.PRICE_NOT_FOUND_DISPLAY)
        self.assertEqual(excel_exporter.effective_price_source("", None, None), "unknown")

    def test_export_latest_sql_uses_latest_snapshot_and_latest_inference(self):
        class FakeCursor:
            def __init__(self):
                self.sql = []
                self.description = [("snapshot_id",)]

            def execute(self, sql, *params):
                self.sql.append(sql)
                return None

            def fetchone(self):
                return [4]

            def fetchall(self):
                return []

        class FakeConn:
            def __init__(self):
                self.cursor_obj = FakeCursor()

            def cursor(self):
                return self.cursor_obj

        conn = FakeConn()
        with mock.patch.object(db_layer, "ensure_monitoring_state_tables", lambda _conn: None), \
             mock.patch.object(db_layer, "_upsert_search", return_value=4):
            db_layer.export_latest_to_rows(conn, "https://example.test/search")
        sql = "\n".join(conn.cursor_obj.sql)
        self.assertIn("ORDER BY ls.SnapshotDate DESC, ls.SnapshotID DESC", sql)
        self.assertIn("latest_price_state", sql)
        self.assertIn("ORDER BY pis.last_attempt_at DESC, pis.updated_at DESC", sql)

    def test_normal_excel_headers_are_exact_and_exclude_debug_columns(self):
        rows = [{"AreaLabel": "Yadboro", "address": "A", "AdPriceDisplay": "$860,000", "InferredPriceRange": "$800,000 - $900,000"}]
        with tempfile.TemporaryDirectory() as tmp, \
             mock.patch.object(excel_exporter, "_connect", return_value=DummyConn()), \
             mock.patch.object(excel_exporter, "_fetch_active_listing_rows", return_value=rows):
            result = excel_exporter.build_active_listings_excel(4, "Yadboro, NSW 2539", output_dir=tmp, mode="normal")
            headers = [cell.value for cell in load_workbook(result.file_path)[excel_exporter.WORKSHEET_NAME][1]]
        self.assertEqual(headers, excel_exporter.NORMAL_HEADERS)
        for header in ["SearchID", "ListingID", "ExternalID", "listing_id", "PriceInferenceLastError", "PriceLow", "PriceHigh", "area_status", "scraped_at", "LandSizeSqm", "BuildingSizeSqm"]:
            self.assertNotIn(header, headers)

    def test_debug_excel_includes_debug_columns(self):
        rows = [{"AreaLabel": "Yadboro", "listing_id": "1", "LandSizeDisplay": "301m²", "LandSizeSqm": 301, "BuildingSizeDisplay": "132m²", "BuildingSizeSqm": 132}]
        with tempfile.TemporaryDirectory() as tmp, \
             mock.patch.object(excel_exporter, "_connect", return_value=DummyConn()), \
             mock.patch.object(excel_exporter, "_fetch_active_listing_rows", return_value=rows):
            result = excel_exporter.build_active_listings_excel(4, "Yadboro, NSW 2539", output_dir=tmp, mode="debug")
            headers = [cell.value for cell in load_workbook(result.file_path)[excel_exporter.WORKSHEET_NAME][1]]
        for header in ["SearchID", "ListingID", "ExternalID", "listing_id", "PriceInferenceStatus", "PriceInferenceLastError", "PriceLow", "PriceHigh", "area_status", "scraped_at", "LandSizeDisplay", "LandSizeSqm", "BuildingSizeDisplay", "BuildingSizeSqm"]:
            self.assertIn(header, headers)

    def test_normal_excel_includes_size_fields_and_price_priority(self):
        rows = [
            {"AreaLabel": "A", "address": "land", "LandSizeDisplay": "8.09ha", "AdPriceDisplay": "$860,000", "InferredPriceRange": "$800,000 - $900,000", "InferredPriceLow": 800000, "InferredPriceHigh": 900000},
            {"AreaLabel": "A", "address": "building", "BuildingSizeDisplay": "132m²", "AdPriceDisplay": "", "InferredPriceRange": "$900,000 - $1,100,000", "InferredPriceLow": 900000, "InferredPriceHigh": 1100000},
            {"AreaLabel": "A", "address": "contact", "AdPriceDisplay": "Contact Agent", "InferredPriceRange": "$900,000 - $1,100,000", "InferredPriceLow": 900000, "InferredPriceHigh": 1100000},
            {"AreaLabel": "A", "address": "35 Mitchell Highway", "LandSizeDisplay": "1,012m²", "AdPriceDisplay": "Call for price", "InferredPriceRange": "", "InferredPriceLow": None, "InferredPriceHigh": None},
        ]
        with tempfile.TemporaryDirectory() as tmp, \
             mock.patch.object(excel_exporter, "_connect", return_value=DummyConn()), \
             mock.patch.object(excel_exporter, "_fetch_active_listing_rows", return_value=rows):
            result = excel_exporter.build_active_listings_excel(4, "Yadboro, NSW 2539", output_dir=tmp, mode="normal")
            sheet = load_workbook(result.file_path)[excel_exporter.WORKSHEET_NAME]
            headers = [cell.value for cell in sheet[1]]
            data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        idx = {name: headers.index(name) for name in headers}
        self.assertEqual(data_rows[0][idx["Land Size"]], "8.09ha")
        self.assertIn(data_rows[0][idx["Building Size"]], (None, ""))
        self.assertEqual(data_rows[1][idx["Building Size"]], "132m²")
        self.assertEqual(data_rows[0][idx["Price"]], "$860,000")
        self.assertEqual(data_rows[0][idx["Ad Price"]], "$860,000")
        self.assertEqual(data_rows[0][idx["Inferred Range"]], "$800,000 - $900,000")
        self.assertEqual(data_rows[1][idx["Price"]], "$900,000 - $1,100,000")
        self.assertEqual(data_rows[2][idx["Price"]], "Contact Agent")
        self.assertEqual(data_rows[2][idx["Inferred Range"]], "$900,000 - $1,100,000")
        self.assertEqual(data_rows[3][idx["Land Size"]], "1,012m²")
        self.assertEqual(data_rows[3][idx["Ad Price"]], "Call for price")
        self.assertEqual(data_rows[3][idx["Price"]], "Call for price")

    def test_normal_excel_excludes_non_active_lifecycle_rows(self):
        rows = [
            {"AreaLabel": "A", "address": "active", "ListingLifecycleStatus": "active", "AdPriceDisplay": "Under Offer"},
            {"AreaLabel": "A", "address": "sold", "ListingLifecycleStatus": "sold", "AdPriceDisplay": "$1"},
            {"AreaLabel": "A", "address": "removed", "ListingLifecycleStatus": "removed", "AdPriceDisplay": "$2"},
            {"AreaLabel": "A", "address": "not_found", "ListingLifecycleStatus": "not_found", "AdPriceDisplay": "$3"},
        ]
        with tempfile.TemporaryDirectory() as tmp, \
             mock.patch.object(excel_exporter, "_connect", return_value=DummyConn()), \
             mock.patch.object(excel_exporter, "_fetch_active_listing_rows", return_value=rows):
            result = excel_exporter.build_active_listings_excel(4, "A", output_dir=tmp, mode="normal")
            sheet = load_workbook(result.file_path)[excel_exporter.WORKSHEET_NAME]
            data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(result.active_listing_count, 1)
        self.assertEqual(data_rows[0][1], "active")

    def test_debug_excel_includes_all_lifecycle_statuses_and_fields(self):
        rows = [
            {"AreaLabel": "A", "address": "active", "ListingLifecycleStatus": "active"},
            {"AreaLabel": "A", "address": "sold", "ListingLifecycleStatus": "sold", "SoldAt": "2026-01-01"},
            {"AreaLabel": "A", "address": "removed", "ListingLifecycleStatus": "removed", "RemovedAt": "2026-01-02"},
        ]
        with tempfile.TemporaryDirectory() as tmp, \
             mock.patch.object(excel_exporter, "_connect", return_value=DummyConn()), \
             mock.patch.object(excel_exporter, "_fetch_active_listing_rows", return_value=rows):
            result = excel_exporter.build_active_listings_excel(4, "A", output_dir=tmp, mode="debug")
            sheet = load_workbook(result.file_path)[excel_exporter.WORKSHEET_NAME]
            headers = [cell.value for cell in sheet[1]]
            data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(data_rows), 3)
        for header in ["ListingLifecycleStatus", "StatusReason", "StatusEvidence", "NotFoundCount", "MissingCount", "FirstNotFoundAt", "LastNotFoundAt", "RemovedAt", "SoldAt", "LastStatusChangeAt"]:
            self.assertIn(header, headers)

    def test_telegram_main_menu_buttons_are_simplified(self):
        self.assertEqual(telegram_bot.MAIN_MENU_BUTTONS, (telegram_bot.BUTTON_ADD, telegram_bot.BUTTON_AREAS, telegram_bot.BUTTON_EXPORT, telegram_bot.BUTTON_HELP))
        source = inspect.getsource(telegram_bot.main_menu_keyboard)
        self.assertNotIn("BUTTON_REMOVE", source)
        self.assertNotIn("BUTTON_CHECK", source)

    def test_my_suburbs_remove_buttons_use_user_area_id_without_export(self):
        source = inspect.getsource(telegram_bot._my_suburbs_keyboard)
        self.assertIn("remove_select:{area['UserAreaID']}", source)
        self.assertNotIn("export", source.lower())
        remove_source = inspect.getsource(telegram_bot.handle_remove_callback)
        self.assertIn("list_user_area_subscriptions", remove_source)
        self.assertIn("remove_user_area_subscription_lifecycle(conn, telegram_user_id, user_area_id)", remove_source)

    def test_help_text_describes_automatic_monitoring_without_check_now(self):
        self.assertIn("Monitoring starts automatically", telegram_bot.HELP_TEXT)
        self.assertIn("Export Excel", telegram_bot.HELP_TEXT)
        self.assertNotIn("Check Now", telegram_bot.HELP_TEXT)
        self.assertNotIn("Check now", telegram_bot.HELP_TEXT)

    def test_baseline_and_detail_refresh_use_same_snapshot_insert_shape(self):
        full_source = inspect.getsource(db_layer.ingest_full_rows)
        detail_source = inspect.getsource(db_layer.ingest_detail_refresh_rows_conn)
        self.assertIn("LISTING_SNAPSHOT_INSERT_SQL", full_source)
        self.assertIn("LISTING_SNAPSHOT_INSERT_SQL", detail_source)
        self.assertIn("merge_enriched_listing_detail_with_latest", detail_source)
        self.assertEqual(db_layer.LISTING_SNAPSHOT_INSERT_SQL.count("?"), 25)

    def test_yadboro_regression_exports_24_inferred_ranges_and_ad_prices(self):
        rows = []
        for idx in range(1, 25):
            has_ad = idx % 2 == 0
            rows.append({
                "listing_id": str(idx),
                "url": f"https://example.test/{idx}",
                "Price": "$850,000" if has_ad else "$800,000 - $850,000",
                "PriceSource": "ad_price" if has_ad else "inferred_range",
                "AdPriceDisplay": "$850,000" if has_ad else "",
                "AdPriceLow": 850000 if has_ad else None,
                "AdPriceHigh": 850000 if has_ad else None,
                "InferredPriceLow": 800000,
                "InferredPriceHigh": 850000,
                "InferredPriceRange": "$800,000 - $850,000",
                "PriceInferenceStatus": "completed",
            })
        with tempfile.TemporaryDirectory() as tmp, \
             mock.patch.object(excel_exporter, "_connect", return_value=DummyConn()), \
             mock.patch.object(excel_exporter, "_fetch_active_listing_rows", return_value=rows):
            result = excel_exporter.build_active_listings_excel(4, "Yadboro, NSW 2539", output_dir=tmp, mode="debug")
            sheet = load_workbook(result.file_path)[excel_exporter.WORKSHEET_NAME]
            headers = [cell.value for cell in sheet[1]]
            data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(data_rows), 24)
        idx = {name: headers.index(name) for name in headers}
        self.assertTrue(all(row[idx["InferredPriceLow"]] for row in data_rows))
        self.assertTrue(all(row[idx["InferredPriceHigh"]] for row in data_rows))
        self.assertTrue(all(row[idx["InferredPriceRange"]] for row in data_rows))
        self.assertTrue(any(row[idx["AdPriceDisplay"]] == "$850,000" for row in data_rows))
        self.assertTrue(any(row[idx["PriceSource"]] == "ad_price" for row in data_rows))
        self.assertTrue(any(row[idx["PriceSource"]] == "inferred_range" for row in data_rows))

    def test_scheduler_ready_active_eligibility(self):
        ready = {"AreaSetupStatus": "ready", "SubscriptionStatus": "active", "SubscriptionNotifyEnabled": 1}
        preparing = {"AreaSetupStatus": "preparing", "SubscriptionStatus": "preparing", "SubscriptionNotifyEnabled": 0}
        paused = {"AreaSetupStatus": "ready", "SubscriptionStatus": "paused", "SubscriptionNotifyEnabled": 0}
        self.assertTrue(monitoring_scheduler._subscription_group_is_ready([ready]))
        self.assertFalse(monitoring_scheduler._subscription_group_is_ready([preparing]))
        self.assertFalse(monitoring_scheduler._subscription_group_is_ready([paused]))

    def test_price_retry_unknowns_skips_when_no_due_rows(self):
        with mock.patch.object(monitoring_scheduler.db_layer, "connect", return_value=DummyConn()), \
             mock.patch.object(monitoring_scheduler, "_search_is_active_for_monitoring", return_value=True), \
             mock.patch.object(monitoring_scheduler.db_layer, "get_due_price_retry_listing_ids", return_value=[]):
            result = monitoring_scheduler.run_price_retry_unknowns_for_search(7, payload={}, dry_run=True)
        self.assertEqual(result["price_retry"]["reason"], "no_due_unknown_prices")

    def test_module2_429_returns_retry_metadata_and_closes_driver(self):
        class FakeDriver:
            def __init__(self):
                self.quit_count = 0

            def quit(self):
                self.quit_count += 1

        fake_driver = FakeDriver()
        with tempfile.TemporaryDirectory() as tmp:
            input_file = os.path.join(tmp, "input.json")
            with open(input_file, "w", encoding="utf-8") as f:
                json.dump([{"listing_id": "1", "url": "u1", "price": "N/A"}], f)
            with mock.patch.object(module2_infer_prices, "build_driver", return_value=fake_driver), \
                 mock.patch.object(module2_infer_prices, "cleanup_chrome_driver"), \
                 mock.patch.object(module2_infer_prices, "infer_prices_window_based_with_checkpoint", return_value=({}, fake_driver, "429")), \
                 mock.patch.object(module2_infer_prices.config, "MODULE2_ROTATE_PROFILE_ON_429", False), \
                 mock.patch("builtins.print"):
                out_csv, out_json = module2_infer_prices.module2_run(
                    "https://www.realestate.com.au/buy/in-test/list-1",
                    input_file=input_file,
                    out_dir=tmp,
                    target_mode="all",
                    target_listing_ids={"1"},
                )
        self.assertIsNone(out_csv)
        self.assertIsNone(out_json)
        self.assertEqual(module2_infer_prices.module2_run.last_result["status"], "429_retry_wait")
        self.assertEqual(fake_driver.quit_count, 1)

    def test_module2_profile_lock_retries_with_unique_profile_once(self):
        class FakeDriver:
            def quit(self):
                pass

        fake_driver = FakeDriver()

        def fake_infer(**kwargs):
            return {
                "1": {
                    "low": 700000,
                    "high": 750000,
                    "window_low": 650000,
                    "window_high": 800000,
                    "found_at_page": 1,
                }
            }, kwargs["driver"], "done"

        with tempfile.TemporaryDirectory() as tmp:
            input_file = os.path.join(tmp, "input.json")
            with open(input_file, "w", encoding="utf-8") as f:
                json.dump([{"listing_id": "1", "url": "u1", "price": "N/A"}], f)
            with mock.patch.object(
                module2_infer_prices,
                "build_driver",
                side_effect=[module2_infer_prices.WebDriverException("Access is denied"), fake_driver],
            ) as build_driver, \
                 mock.patch.object(module2_infer_prices, "cleanup_chrome_driver"), \
                 mock.patch.object(module2_infer_prices, "infer_prices_window_based_with_checkpoint", side_effect=fake_infer), \
                 mock.patch("builtins.print"):
                out_csv, out_json = module2_infer_prices.module2_run(
                    "https://www.realestate.com.au/buy/in-test/list-1",
                    input_file=input_file,
                    out_dir=tmp,
                    target_mode="all",
                    target_listing_ids={"1"},
                    checkpoint_search_id=22,
                )

        self.assertIsNotNone(out_json)
        self.assertEqual(build_driver.call_count, 2)
        first_profile = build_driver.call_args_list[0].kwargs["profile_dir_override"]
        second_profile = build_driver.call_args_list[1].kwargs["profile_dir_override"]
        self.assertNotEqual(first_profile, second_profile)
        self.assertIn("module2_chrome_profiles", second_profile)

    def test_scheduler_does_not_enqueue_duplicate_price_refresh_when_active_exists(self):
        now = monitoring_scheduler._utcnow()
        job_queue.enable_in_memory_store()
        try:
            active = job_queue.enqueue_job_once(
                job_queue.JOB_TYPE_PRICE_REFRESH_EXISTING,
                search_id=2,
                priority=job_queue.PRIORITY_PRICE_REFRESH,
                run_after=now,
                payload={"run_id": "existing"},
            )
            job_queue.claim_next_job("test-worker")
            subscriptions = [{
                "SearchID": 2,
                "UserAreaID": 10,
                "SearchURL": "https://example.test/search",
                "AreaSetupStatus": "ready",
                "AreaReadyAt": now - monitoring_scheduler.timedelta(days=1),
                "SubscriptionStatus": "active",
                "SubscriptionNotifyEnabled": 1,
                "LastLightCheckAt": now,
                "LastDetailRefreshAt": now,
                "LastPriceRefreshAt": None,
                "LastFullListingSweepAt": now,
            }]
            with mock.patch.object(monitoring_scheduler.db_layer, "connect", return_value=DummyConn()), \
                 mock.patch.object(monitoring_scheduler.db_layer, "ensure_telegram_bot_tables", lambda _conn: None), \
                 mock.patch.object(monitoring_scheduler.db_layer, "get_active_user_area_subscriptions", return_value=subscriptions), \
                 mock.patch.object(monitoring_scheduler.db_layer, "get_due_price_retry_listing_ids", return_value=[]), \
                 mock.patch.object(monitoring_scheduler, "_fetch_sql_server_local_time", return_value=now):
                result = monitoring_scheduler.enqueue_due_monitoring_jobs(now=now)

            price_jobs = [
                row for row in job_queue.get_active_jobs()
                if int(row.get("SearchID") or 0) == 2
                and row.get("JobType") in {job_queue.JOB_TYPE_PRICE_REFRESH_EXISTING, job_queue.JOB_TYPE_MODULE2_PRICE_REFRESH_AREA}
            ]
            self.assertEqual(len(price_jobs), 1)
            self.assertEqual(price_jobs[0]["JobID"], active["JobID"])
            self.assertTrue(any(item.get("reason") == "active_price_refresh_exists" for item in result["skipped_duplicates"]))
        finally:
            job_queue.disable_in_memory_store()

    def test_remove_area_with_another_active_subscriber_keeps_search_active(self):
        conn = LifecycleConn([
            {"UserAreaID": 42, "TelegramUserID": 1, "SearchID": 7, "IsActive": 1, "SubscriptionStatus": "active", "NotifyEnabled": 1},
            {"UserAreaID": 99, "TelegramUserID": 2, "SearchID": 7, "IsActive": 1, "SubscriptionStatus": "active", "NotifyEnabled": 1},
        ], jobs=[{"JobID": 1, "SearchID": 7, "Status": "queued"}], area_states={7: {"setup_status": "ready"}})
        with mock.patch.object(db_layer, "ensure_telegram_bot_tables", lambda _conn: None), \
             mock.patch.object(db_layer, "ensure_monitoring_state_tables", lambda _conn: None):
            result = db_layer.remove_user_area_subscription_lifecycle(conn, 1, 42)

        self.assertTrue(result["removed"])
        self.assertEqual(result["resolved_search_id"], 7)
        self.assertEqual(result["resolved_area_id"], 7)
        self.assertEqual(result["remaining_active_subscriptions"], 1)
        self.assertEqual(result["action"], "kept_active")
        self.assertEqual(conn.area_states[7]["setup_status"], "ready")
        self.assertEqual(conn.jobs[0]["Status"], "queued")
        self.assertEqual(conn.subscriptions[1]["SubscriptionStatus"], "active")

    def test_remove_last_subscriber_inactivates_correct_search_id_and_cancels_jobs(self):
        conn = LifecycleConn([
            {"UserAreaID": 42, "TelegramUserID": 1, "SearchID": 7, "IsActive": 1, "SubscriptionStatus": "active", "NotifyEnabled": 1},
        ], jobs=[
            {"JobID": 1, "SearchID": 7, "Status": "queued"},
            {"JobID": 2, "SearchID": 42, "Status": "queued"},
        ], area_states={7: {"setup_status": "ready"}, 42: {"setup_status": "ready"}})
        with mock.patch.object(db_layer, "ensure_telegram_bot_tables", lambda _conn: None), \
             mock.patch.object(db_layer, "ensure_monitoring_state_tables", lambda _conn: None):
            result = db_layer.remove_user_area_subscription_lifecycle(conn, 1, 42)

        self.assertTrue(result["removed"])
        self.assertEqual(result["resolved_search_id"], 7)
        self.assertEqual(result["remaining_active_subscriptions"], 0)
        self.assertEqual(result["action"], "inactivated")
        self.assertEqual(conn.area_states[7]["setup_status"], "inactive")
        self.assertEqual(conn.jobs[0]["Status"], "cancelled")
        self.assertEqual(conn.jobs[1]["Status"], "queued")

    def test_worker_skips_inactive_area_job_before_browser(self):
        now = monitoring_scheduler._utcnow()
        job_queue.enable_in_memory_store([{
            "JobID": 1,
            "JobType": job_queue.JOB_TYPE_BASELINE_SETUP_AREA,
            "SearchID": 7,
            "UserAreaID": 42,
            "Priority": job_queue.PRIORITY_SETUP,
            "Status": "queued",
            "RunAfter": now,
            "AttemptCount": 0,
            "MaxAttempts": 3,
            "PayloadJson": json.dumps({"search_url": "https://example.test/search"}),
            "DedupeKey": "baseline_setup_area:area_id=7",
            "CreatedAt": now,
            "UpdatedAt": now,
        }])
        try:
            with mock.patch.object(monitoring_scheduler, "_search_is_active_for_monitoring", return_value=False), \
                 mock.patch.object(monitoring_scheduler, "baseline_setup_area") as baseline_mock:
                result = monitoring_scheduler.run_next_job_once(worker_id="test-worker")
            self.assertEqual(result["job_result"]["status"], "cancelled")
            baseline_mock.assert_not_called()
            jobs = job_queue.get_active_jobs()
            self.assertEqual(jobs, [])
            failed_or_cancelled = [row for row in job_queue._TEST_STORE if row["JobID"] == 1]
            self.assertEqual(failed_or_cancelled[0]["Status"], "cancelled")
        finally:
            job_queue.disable_in_memory_store()

    def test_worker_keeps_realestate_rate_limit_as_retry_wait_without_consuming_attempt(self):
        now = monitoring_scheduler._utcnow()
        job_queue.enable_in_memory_store([{
            "JobID": 1,
            "JobType": job_queue.JOB_TYPE_BASELINE_SETUP_AREA,
            "SearchID": 7,
            "UserAreaID": 42,
            "Priority": job_queue.PRIORITY_SETUP,
            "Status": "queued",
            "RunAfter": now,
            "AttemptCount": 2,
            "MaxAttempts": 3,
            "PayloadJson": json.dumps({"search_url": "https://example.test/search"}),
            "DedupeKey": "baseline_setup_area:area_id=7",
            "CreatedAt": now,
            "UpdatedAt": now,
        }])
        try:
            with mock.patch.object(monitoring_scheduler, "_search_is_active_for_monitoring", return_value=True), \
                 mock.patch.object(monitoring_scheduler, "baseline_setup_area", side_effect=RealEstateRateLimitedError("realestate_rate_limited_or_blocked_http_429", retry_after_seconds=21600)), \
                 mock.patch.object(monitoring_scheduler.db_layer, "connect", return_value=DummyConn()), \
                 mock.patch.object(monitoring_scheduler.db_layer, "upsert_area_monitoring_state"):
                result = monitoring_scheduler.run_next_job_once(worker_id="test-worker")
            self.assertEqual(result["status"], "completed")
            row = next(item for item in job_queue._TEST_STORE if item["JobID"] == 1)
            self.assertEqual(row["Status"], "retry_wait")
            self.assertEqual(row["AttemptCount"], 2)
            self.assertEqual(row["LastError"], "realestate_rate_limited_or_blocked_http_429")
        finally:
            job_queue.disable_in_memory_store()

    def test_blocked_light_check_does_not_mark_search_checked(self):
        job = {"JobType": job_queue.JOB_TYPE_LIGHT_CHECK_NEW_LISTINGS, "SearchID": 7, "UserAreaID": 42}
        sub = {"UserAreaID": 42, "SearchID": 7, "SearchURL": "https://example.test/search"}
        blocked = {"scan_status": "blocked_rate_limited", "blocked_reason": "blocked_kpsdk", "stop_reason": "blocked_kpsdk", "trusted_scan": False}
        with mock.patch.object(monitoring_scheduler, "_load_search_subscription", return_value=sub), \
             mock.patch.object(monitoring_scheduler, "_search_is_active_for_monitoring", return_value=True), \
             mock.patch.object(monitoring_scheduler, "light_check_area", return_value=blocked), \
             mock.patch.object(monitoring_scheduler.db_layer, "mark_search_light_checked") as mark_checked:
            result = monitoring_scheduler.execute_job(job, send_telegram=False)

        self.assertEqual(result["status"], "retry_wait")
        self.assertEqual(result["reason"], "blocked_kpsdk")
        mark_checked.assert_not_called()

    def test_scheduler_does_not_count_inactive_search_as_ready(self):
        now = monitoring_scheduler._utcnow()
        job_queue.enable_in_memory_store()
        try:
            with mock.patch.object(monitoring_scheduler.db_layer, "connect", return_value=DummyConn()), \
                 mock.patch.object(monitoring_scheduler.db_layer, "ensure_telegram_bot_tables", lambda _conn: None), \
                 mock.patch.object(monitoring_scheduler.db_layer, "get_active_user_area_subscriptions", return_value=[]), \
                 mock.patch.object(monitoring_scheduler, "_fetch_sql_server_local_time", return_value=now):
                result = monitoring_scheduler.enqueue_due_monitoring_jobs(now=now)
            self.assertEqual(result["ready_search_ids_considered"], [])
            self.assertEqual(result["not_ready_search_ids_considered"], [])
            self.assertEqual(result["created"], [])
        finally:
            job_queue.disable_in_memory_store()

    def test_reactivation_baseline_dedupe_enqueues_once_for_inactive_area(self):
        conn = DummyConn()
        job_queue.enable_in_memory_store()
        try:
            with mock.patch.object(db_layer, "ensure_monitoring_state_tables", lambda _conn: None), \
                 mock.patch.object(db_layer, "get_area_monitoring_state", return_value={"setup_status": "inactive"}), \
                 mock.patch.object(db_layer, "upsert_area_monitoring_state"):
                first = db_layer.enqueue_baseline_setup_job(conn, 7, "https://example.test/search")
                second = db_layer.enqueue_baseline_setup_job(conn, 7, "https://example.test/search")
            self.assertTrue(first["created"])
            self.assertFalse(second["created"])
            self.assertTrue(second["duplicate"])
        finally:
            job_queue.disable_in_memory_store()

    def test_add_active_ready_area_does_not_enqueue_baseline(self):
        with mock.patch.object(db_layer, "ensure_telegram_bot_tables", lambda _conn: None), \
             mock.patch.object(db_layer, "ensure_sort_list_date", lambda url: url), \
             mock.patch.object(db_layer, "get_or_create_suburb_search", return_value=(7, False)), \
             mock.patch.object(db_layer, "user_already_subscribed", return_value=False), \
             mock.patch.object(db_layer, "active_area_count_for_user", return_value=0), \
             mock.patch.object(db_layer, "get_search_setup_state", return_value={"state": "ready", "is_ready": True, "baseline_completed": True, "detail_started": True, "price_completed": True, "detail_baseline_status": "completed"}), \
             mock.patch.object(db_layer, "create_or_reactivate_subscription", return_value=("created", {"user_area_id": 42, "search_id": 7})), \
             mock.patch.object(db_layer, "upsert_user_area_subscription_state") as substate_mock, \
             mock.patch.object(db_layer, "enqueue_baseline_setup_job") as enqueue_mock:
            ok, payload = db_layer.add_user_area_subscription(DummyConn(), 1, "url", "Area")
        self.assertTrue(ok)
        self.assertEqual(payload["reason"], "ready")
        self.assertEqual(payload["message"], "Monitoring is already active for this area.")
        enqueue_mock.assert_not_called()
        substate_mock.assert_called_once()

    def test_heartbeat_separates_inactive_failed_jobs(self):
        now = monitoring_scheduler._utcnow()
        job_queue.enable_in_memory_store([
            {"JobID": 1, "JobType": job_queue.JOB_TYPE_SETUP_PRICE_BASELINE, "SearchID": 7, "Status": "failed", "AttemptCount": 1, "LastError": "active failure", "UpdatedAt": now, "AreaActive": True},
            {"JobID": 2, "JobType": job_queue.JOB_TYPE_SETUP_PRICE_BASELINE, "SearchID": 8, "Status": "failed", "AttemptCount": 1, "LastError": "inactive failure", "UpdatedAt": now, "AreaActive": False},
        ])
        try:
            summary = job_queue.get_failed_job_summary_by_lifecycle(limit=5)
        finally:
            job_queue.disable_in_memory_store()
        self.assertEqual([row["job_id"] for row in summary["active_failed_jobs"]], [1])
        self.assertEqual([row["job_id"] for row in summary["inactive_failed_jobs"]], [2])

    def test_long_job_error_is_not_truncated_app_side(self):
        now = monitoring_scheduler._utcnow()
        long_error = "x" * 5000
        job_queue.enable_in_memory_store([{
            "JobID": 1,
            "JobType": job_queue.JOB_TYPE_SETUP_PRICE_BASELINE,
            "SearchID": 7,
            "Status": "running",
            "Priority": job_queue.PRIORITY_SETUP,
            "RunAfter": now,
            "AttemptCount": 0,
            "MaxAttempts": 1,
            "CreatedAt": now,
            "UpdatedAt": now,
        }])
        try:
            row = job_queue.mark_job_failed(1, long_error, retryable=False)
        finally:
            job_queue.disable_in_memory_store()
        self.assertEqual(len(row["LastError"]), 5000)

    def test_reactivated_area_runs_full_baseline_modules(self):
        rows = [{"listing_id": "1", "url": "u1", "price": "N/A"}]
        outcome = self._run_mocked_baseline(rows, [dict(rows[0], address="A")])
        self.assertEqual(outcome["calls"], ["module1", "module3", "module2"])
        self.assertEqual(outcome["module2_kwargs"][0]["target_mode"], "all")
        self.assertEqual(outcome["result"]["status"], "ready")


class TelegramExportAndReadyTests(unittest.IsolatedAsyncioTestCase):
    def _message_update(self):
        message = types.SimpleNamespace(
            reply_photo=mock.AsyncMock(),
            reply_text=mock.AsyncMock(),
        )
        return types.SimpleNamespace(message=message)

    def _export_update(self, data="export_area:4"):
        query = types.SimpleNamespace(
            data=data,
            answer=mock.AsyncMock(),
            edit_message_text=mock.AsyncMock(),
        )
        chat = types.SimpleNamespace(
            send_document=mock.AsyncMock(),
            send_message=mock.AsyncMock(),
        )
        return types.SimpleNamespace(callback_query=query, effective_chat=chat)

    async def test_start_sends_image_and_caption_in_one_message(self):
        update = self._message_update()
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
            f.write(b"png")
            image_path = f.name
        try:
            with mock.patch.object(telegram_bot, "START_IMAGE_PATH", image_path), \
                 mock.patch.object(telegram_bot, "register_chat", return_value=1), \
                 mock.patch.object(telegram_bot, "_clear_session"), \
                 mock.patch.object(telegram_bot, "main_menu_keyboard", return_value=None):
                await telegram_bot.start(update, types.SimpleNamespace())
        finally:
            os.remove(image_path)

        update.message.reply_photo.assert_awaited_once()
        update.message.reply_text.assert_not_awaited()
        kwargs = update.message.reply_photo.await_args.kwargs
        self.assertIn("Welcome to OzHome Monitor", kwargs["caption"])
        self.assertIn("Choose an option below to get started", kwargs["caption"])

    def test_start_image_path_defaults_to_project_assets(self):
        self.assertEqual(os.path.basename(telegram_bot.START_IMAGE_PATH), "start.png")
        self.assertEqual(os.path.basename(os.path.dirname(telegram_bot.START_IMAGE_PATH)), "assets")
        self.assertTrue(os.path.exists(telegram_bot.START_IMAGE_PATH))

    async def test_area_candidate_callback_uses_user_session_payload(self):
        area = {
            "label": "Petersham, NSW 2049",
            "search_url": "https://example.test/petersham",
            "suburb_name": "Petersham",
            "postcode": "2049",
        }
        query = types.SimpleNamespace(
            data="area_candidate:0",
            answer=mock.AsyncMock(),
            edit_message_text=mock.AsyncMock(),
        )
        update = types.SimpleNamespace(
            callback_query=query,
            effective_chat=types.SimpleNamespace(send_message=mock.AsyncMock()),
        )

        with mock.patch.object(telegram_bot, "register_chat", return_value=1), \
             mock.patch.object(telegram_bot, "_session", return_value={"state": "choosing_area_candidate", "payload": {"matches": [area]}}), \
             mock.patch.object(telegram_bot, "_set_session") as set_session_mock, \
             mock.patch.object(telegram_bot, "_confirm_area_keyboard", return_value=None):
            await telegram_bot.handle_area_callback(update, types.SimpleNamespace())

        set_session_mock.assert_called_once_with(1, "confirming_area", {"pending_area": area})
        query.edit_message_text.assert_awaited_once()

    async def test_telegram_export_handler_fallback_without_area_label(self):
        update = self._export_update()
        area = {"UserAreaID": 4, "SearchID": 4, "name": "Yadboro, NSW 2539"}
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(b"placeholder")
            file_path = f.name
        result = types.SimpleNamespace(file_path=file_path, active_listing_count=24, generated_at=datetime(2026, 1, 1, 12, 0, 0))
        try:
            with mock.patch.object(telegram_bot, "register_chat", return_value=1), \
                 mock.patch.object(telegram_bot.excel_exporter, "get_authorized_export_area", return_value=area), \
                 mock.patch.object(telegram_bot.excel_exporter, "build_active_listings_excel", return_value=result) as export_mock, \
                 mock.patch.object(telegram_bot, "main_menu_keyboard", return_value=None):
                await telegram_bot.handle_export_area_selection(update, types.SimpleNamespace())
        finally:
            os.remove(file_path)

        update.effective_chat.send_document.assert_awaited_once()
        self.assertEqual(export_mock.call_args.kwargs["mode"], "normal")
        caption = update.effective_chat.send_document.await_args.kwargs["caption"]
        self.assertIn("Area: Yadboro, NSW 2539", caption)

    async def test_excel_export_handler_sends_file(self):
        update = self._export_update()
        area = {"UserAreaID": 4, "SearchID": 4, "AreaLabel": "Yadboro, NSW 2539"}
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(b"placeholder")
            file_path = f.name
        result = excel_exporter.ExcelExportResult(file_path=file_path, generated_at=datetime(2026, 1, 1, 12, 0, 0), active_listing_count=24, area_label="Yadboro, NSW 2539", search_id=4)
        try:
            with mock.patch.object(telegram_bot, "register_chat", return_value=1), \
                 mock.patch.object(telegram_bot.excel_exporter, "get_authorized_export_area", return_value=area), \
                 mock.patch.object(telegram_bot.excel_exporter, "build_active_listings_excel", return_value=result) as export_mock, \
                 mock.patch.object(telegram_bot, "main_menu_keyboard", return_value=None):
                await telegram_bot.handle_export_area_selection(update, types.SimpleNamespace())
        finally:
            os.remove(file_path)

        update.effective_chat.send_document.assert_awaited_once()
        self.assertEqual(export_mock.call_args.kwargs["mode"], "normal")
        self.assertEqual(update.effective_chat.send_document.await_args.kwargs["filename"], result.filename)

    async def test_admin_debug_export_uses_debug_mode_when_configured(self):
        update = self._export_update()
        area = {"UserAreaID": 4, "SearchID": 4, "AreaLabel": "Yadboro, NSW 2539"}
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(b"placeholder")
            file_path = f.name
        result = excel_exporter.ExcelExportResult(file_path=file_path, generated_at=datetime(2026, 1, 1, 12, 0, 0), active_listing_count=1, area_label="Yadboro, NSW 2539", search_id=4)
        try:
            with mock.patch.object(telegram_bot.config, "EXCEL_EXPORT_MODE", "debug"), \
                 mock.patch.object(telegram_bot, "_is_admin_chat", return_value=True), \
                 mock.patch.object(telegram_bot, "register_chat", return_value=1), \
                 mock.patch.object(telegram_bot.excel_exporter, "get_authorized_export_area", return_value=area), \
                 mock.patch.object(telegram_bot.excel_exporter, "build_active_listings_excel", return_value=result) as export_mock, \
                 mock.patch.object(telegram_bot, "main_menu_keyboard", return_value=None):
                await telegram_bot.handle_export_area_selection(update, types.SimpleNamespace())
        finally:
            os.remove(file_path)
        self.assertEqual(export_mock.call_args.kwargs["mode"], "debug")

    async def test_export_failure_includes_debug_id_and_logs_traceback(self):
        update = self._export_update()
        area = {"UserAreaID": 4, "SearchID": 4, "AreaLabel": "Yadboro, NSW 2539"}
        with mock.patch.object(telegram_bot, "register_chat", return_value=1), \
             mock.patch.object(telegram_bot.excel_exporter, "get_authorized_export_area", return_value=area), \
             mock.patch.object(telegram_bot.excel_exporter, "build_active_listings_excel", side_effect=RuntimeError("boom")), \
             mock.patch.object(telegram_bot, "main_menu_keyboard", return_value=None), \
             self.assertLogs(telegram_bot.logger, level="ERROR") as logs:
            await telegram_bot.handle_export_area_selection(update, types.SimpleNamespace())

        message = update.effective_chat.send_message.await_args.args[0]
        self.assertIn("Debug ID: EXPORT-", message)
        debug_id = message.rsplit("Debug ID: ", 1)[1]
        self.assertTrue(any(debug_id in line and "Traceback" in line for line in logs.output))

    def test_baseline_ready_notification_activates_and_sends_summary(self):
        job = {
            "JobType": job_queue.JOB_TYPE_BASELINE_SETUP_AREA,
            "SearchID": 4,
            "UserAreaID": 4,
            "PayloadJson": json.dumps({"search_url": "https://example.test/search"}),
        }
        subscription = {
            "UserAreaID": 4,
            "TelegramUserID": 1,
            "ChatID": "1",
            "SearchID": 4,
            "AreaLabel": "Yadboro, NSW 2539",
            "SubscriptionStatus": "preparing",
            "SubscriptionNotifyEnabled": 0,
            "ReadySummarySentAt": None,
        }

        def activate(_conn, area_id):
            self.assertEqual(area_id, 4)
            subscription["SubscriptionStatus"] = "active"
            subscription["SubscriptionNotifyEnabled"] = 1

        sent_texts = []

        def send_once(sub, summary_type):
            sent_texts.append(monitoring_scheduler.SETUP_SUMMARY_TEXT[summary_type](sub, None))
            return {"status": "sent", "user_area_id": sub["UserAreaID"], "summary_type": summary_type}

        with mock.patch.object(monitoring_scheduler, "baseline_setup_area", return_value={"status": "ready", "rows_full": 24, "inferred_price_count": 24, "unknown_price_count": 0}), \
             mock.patch.object(monitoring_scheduler, "_search_is_active_for_monitoring", return_value=True), \
             mock.patch.object(monitoring_scheduler.db_layer, "connect", return_value=DummyConn()), \
             mock.patch.object(monitoring_scheduler.db_layer, "activate_area_subscriptions", side_effect=activate) as activate_mock, \
             mock.patch.object(monitoring_scheduler.db_layer, "get_active_user_area_subscriptions_for_search", return_value=[subscription]), \
             mock.patch.object(monitoring_scheduler, "_send_setup_summary_once", side_effect=send_once):
            result = monitoring_scheduler.execute_job(job, send_telegram=True)

        activate_mock.assert_called_once()
        self.assertEqual(subscription["SubscriptionStatus"], "active")
        self.assertEqual(subscription["SubscriptionNotifyEnabled"], 1)
        self.assertEqual(result["status"], "ready")
        self.assertIn("✅ Monitoring is active", sent_texts[0])
        self.assertIn("Yadboro, NSW 2539", sent_texts[0])
        self.assertIn("I found 24 active listings.", sent_texts[0])
        self.assertIn("Price ranges were inferred for 24 listings.", sent_texts[0])
        self.assertIn("All listing price ranges were inferred successfully.", sent_texts[0])


if __name__ == "__main__":
    unittest.main()
