import os
import json
from typing import Optional

from config import (
    AREA_SEARCH_URL,
    DB_PATH,
    MAX_PAGES_MODULE1,
    MODULE2_MAX_HIGH,
    MODULE2_MAX_PAGES_PER_WINDOW,
    MODULE2_STEP,
    MODULE2_WINDOW_WIDTH,
    MODULE3_EMPTY_RETRY,
    MODULE3_SLEEP_BETWEEN,
    MODULE3_WAIT_TIMEOUT,
    OUTPUT_DIR,
    PIPELINE_TIMEOUT,
)

# سه ماژول شما
import module1_list_scraper
import module2_infer_prices
import module3_enrich_details

from db_layer import init_db, ingest_full_rows, connect, export_latest_to_rows

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def export_to_excel_from_db(db_path: str, area_url: str, output_dir: str) -> str:
    os.makedirs(output_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"

    conn = connect(db_path)
    rows = export_latest_to_rows(conn, area_url)
    conn.close()

    headers = [
        "listing_id", "url", "address", "property_type", "bedrooms", "bathrooms", "parking",
        "price_display", "price_low", "price_high", "price_method",
        "inspection_short", "inspection_long", "auction_label", "auction_time",
        "agency_name", "agency_code", "agency_profile_url", "agency_address",
        "agents",
        "detail_price_display",
        "area_status", "area_last_seen_at", "scraped_at",
        "description"
    ]

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    width_map = {
        "listing_id": 12,
        "url": 40,
        "address": 32,
        "property_type": 14,
        "price_display": 18,
        "agency_name": 22,
        "agents": 35,
        "description": 60,
    }
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width_map.get(h, 16)

    ts = __import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(output_dir, f"export_{ts}.xlsx")
    wb.save(out_path)
    return out_path


def run_pipeline_for_area(area_url: str, max_pages: Optional[int] = None, timeout: int = 25) -> Optional[str]:
    """
    max_pages:
      - None => تا آخرین صفحه لیست می‌ره
      - عدد  => برای تست محدود می‌کنه
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 0) DB init
    init_db(DB_PATH)

    # 1) Module 1 (✅ حالا تا آخر می‌ره چون max_pages=None)
    rows1 = module1_list_scraper.scrape_search(area_url, max_pages=max_pages, timeout=timeout)
    csv1, json1 = module1_list_scraper.save_results(rows1, out_dir=OUTPUT_DIR)

    # 2) Module 2
    csv2, json2 = module2_infer_prices.module2_run(
        base_list_url=area_url,
        input_file=json1,
        out_dir=OUTPUT_DIR,
        window_width=MODULE2_WINDOW_WIDTH,
        step=MODULE2_STEP,
        max_high=MODULE2_MAX_HIGH,
        max_pages_per_window=MODULE2_MAX_PAGES_PER_WINDOW,
        only_overwrite_na=True,
        smart_start=True,
    )
    if not json2:
        return None

    # 3) Module 3
    csv3, json3 = module3_enrich_details.module3_run(
        area_search_url=area_url,
        input_file=json2,
        out_dir=OUTPUT_DIR,
        only_if_missing=True,
        wait_timeout=MODULE3_WAIT_TIMEOUT,
        sleep_between=MODULE3_SLEEP_BETWEEN,
        empty_retry=MODULE3_EMPTY_RETRY,
    )
    if not json3:
        return None

    # 4) Ingest to DB
    with open(json3, "r", encoding="utf-8") as f:
        full_rows = json.load(f)

    run_id = ingest_full_rows(DB_PATH, area_url, full_rows)
    print(f"✅ DB ingested. run_id={run_id}")

    # 5) Export Excel from DB
    xlsx_path = export_to_excel_from_db(DB_PATH, area_url, OUTPUT_DIR)
    print(f"✅ Excel exported: {xlsx_path}")
    return xlsx_path


if __name__ == "__main__":
    # ✅ تا آخرین صفحه:
    run_pipeline_for_area(AREA_SEARCH_URL, max_pages=MAX_PAGES_MODULE1, timeout=PIPELINE_TIMEOUT)

    # اگر خواستی برای تست فقط 3 صفحه:
    # run_pipeline_for_area(AREA_SEARCH_URL, max_pages=3, timeout=25)
